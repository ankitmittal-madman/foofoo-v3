from pathlib import Path
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = Path("indian_meal_habits_template.xlsx").resolve()

USER = "user_entered"
DERIVED = "derived"
AI = "ai_assisted_normalized"
SEED = "externally_seeded_master"
USAGE = "usage_generated"
HYBRID = "hybrid_seed_ai_usage"

def f(name, question, dtype="text", allowed="free_text", rule="non-empty text", source=USER, req="no", ex="", notes=""):
    return dict(field_name=name, user_question=question, data_type=dtype, allowed_values=allowed,
                validation_rule=rule, source_category=source, required_yes_no=req,
                example_value=ex, notes=notes)

schemas = {
"households": [
 f("household_id","System household identifier","uuid","UUID","unique; immutable",DERIVED,"yes","HH-0001","Primary key"),
 f("household_name","What should we call your household?","varchar(80)","free_text","1-80 chars",USER,"yes","Sharma Home"),
 f("current_city_id","Which city do you live in now?","varchar(20)","city_master.city_id","must exist in city_master",USER,"yes","CITY_DELHI"),
 f("current_state_id","Current state/UT","varchar(20)","state_master.state_id","must exist in state_master",DERIVED,"yes","ST_DL","Derived from city"),
 f("origin_state_ids","Which states/regions influence your home food?","json array","state_master.state_id[]","1-5 unique IDs",USER,"no",'["ST_PB","ST_DL"]'),
 f("living_setup","What best describes your household?","enum","single|shared_flat|couple|nuclear_family|joint_family|hostel_pg|multigenerational","one allowed value",USER,"yes","nuclear_family"),
 f("household_size","How many people usually eat at home?","smallint","1-30","integer 1-30",USER,"yes",4),
 f("adult_count","Adults (18-59)?","smallint","0-30","integer >=0; counts total to household_size",USER,"yes",2),
 f("child_count","Children (2-17)?","smallint","0-20","integer >=0",USER,"yes",1),
 f("elder_count","Elders (60+)?","smallint","0-20","integer >=0",USER,"yes",1),
 f("infant_count","Infants/toddlers (0-1)?","smallint","0-10","integer >=0",USER,"yes",0),
 f("decision_model","Who usually decides meals?","enum","self|partner|cook|elder|shared|rotating","one allowed value",USER,"yes","shared"),
 f("monthly_food_budget_inr","Approximate monthly household food budget?","decimal(10,2)","INR","0-1000000",USER,"no",18000),
 f("created_at","Created timestamp","timestamp","ISO-8601 UTC","valid timestamp",DERIVED,"yes","2026-07-01T09:00:00Z")],
"users": [
 f("user_id","System user identifier","uuid","UUID","unique; immutable",DERIVED,"yes","U-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"),
 f("display_name","What should we call you?","varchar(80)","free_text","1-80 chars",USER,"yes","Asha"),
 f("age_band","Your age range?","enum","13_17|18_24|25_34|35_44|45_59|60_74|75_plus|prefer_not_to_say","one value",USER,"yes","35_44"),
 f("gender","Gender (optional)","enum","woman|man|non_binary|self_describe|prefer_not_to_say","one value",USER,"no","woman","Collect only if product has a tested use"),
 f("persona_codes","Which descriptions fit you?","json array","persona_master.persona_code[]","1-6 unique codes",USER,"yes",'["working_professional","taste_first"]'),
 f("is_primary_user","Are you setting up this household?","boolean","true|false","boolean",DERIVED,"yes",True),
 f("onboarding_depth","Onboarding path","enum","mvp|full","one value",DERIVED,"yes","mvp"),
 f("consent_personalization","Allow preferences to personalize recommendations?","boolean","true|false","explicit consent required",USER,"yes",True),
 f("profile_updated_at","Last profile update","timestamp","ISO-8601 UTC","valid timestamp",DERIVED,"yes","2026-07-01T09:05:00Z")],
"members": [
 f("member_id","Member identifier","uuid","UUID","unique; immutable",DERIVED,"yes","M-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"),
 f("linked_user_id","Linked app user, if any","uuid nullable","users.user_id","valid FK or blank",DERIVED,"no","U-0001"),
 f("member_label","Name or nickname","varchar(80)","free_text","1-80 chars",USER,"yes","Asha"),
 f("life_stage","Life stage","enum","infant|child|adult|elder|pregnant|postpartum|recovery","one value",USER,"yes","adult"),
 f("meal_participation","Which home meals do they usually eat?","json array","breakfast|lunch|dinner|snack","1-4 unique values",USER,"yes",'["breakfast","dinner"]'),
 f("is_decision_maker","Do they help decide meals?","boolean","true|false","boolean",USER,"yes",True),
 f("notes","Anything essential for planning?","varchar(250)","free_text","max 250 chars; no diagnosis detail",USER,"no","Carries lunch to office")],
"food_preferences": [
 f("preference_id","Preference record ID","uuid","UUID","unique",DERIVED,"yes","FP-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"), f("member_id","Member or household-level blank","uuid nullable","members.member_id","valid FK or blank",USER,"no","M-0001"),
 f("diet_pattern","Main eating pattern","enum","vegetarian|eggetarian|non_vegetarian|vegan|jain","one value",USER,"yes","vegetarian"),
 f("fasting_patterns","Do you observe food-related fasts?","json array","none|weekly|navratri|ekadashi|ramadan|lent|other","unique values",USER,"no",'["navratri"]'),
 f("spice_level","Preferred heat level","tinyint","1-5","integer 1-5",USER,"yes",3), f("sweetness_level","Preferred sweetness","tinyint","1-5","integer 1-5",USER,"no",2), f("oil_level","Preferred oil richness","tinyint","1-5","integer 1-5",USER,"yes",2),
 f("texture_preference","Gravy or dry dishes?","enum","mostly_dry|balanced|mostly_gravy|no_preference","one value",USER,"no","balanced"),
 f("home_cooked_days_week","How many days per week is food home-cooked?","tinyint","0-7","integer 0-7",USER,"yes",6),
 f("eat_out_meals_week","Meals eaten out per week?","tinyint","0-21","integer 0-21",USER,"no",1), f("delivery_meals_week","Meals ordered per week?","tinyint","0-21","integer 0-21",USER,"no",1),
 f("comfort_dish_ids","Your comfort foods","json array","dish_master.dish_id[]","max 10 unique IDs",AI,"no",'["DISH_RAJMA_CHWAL","DISH_KHICHDI"]',"Confirm AI alias match")],
"regional_taste": [
 f("regional_taste_id","Regional taste record ID","uuid","UUID","unique",DERIVED,"yes","RT-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"),
 f("cuisine_region_id","Which regional cuisine?","varchar(30)","cuisine_region_master.region_id","valid FK",USER,"yes","REG_PUNJABI"),
 f("preference_score","How much does your household enjoy it?","tinyint","1-5","integer 1-5",USER,"yes",5),
 f("authenticity_preference","How traditional should recommendations be?","enum","flexible|balanced|regional_purist","one value",USER,"no","balanced"),
 f("familiarity","Experience with this cuisine","enum","unfamiliar|somewhat_familiar|regular|heritage","one value",USER,"yes","heritage"),
 f("source_reason","Why is this cuisine relevant?","enum","origin|current_location|partner_family|preference|learned","one value",USER,"no","origin")],
"exclusions": [
 f("exclusion_id","Exclusion ID","uuid","UUID","unique",DERIVED,"yes","EX-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"), f("member_id","Affected member; blank means whole home","uuid nullable","members.member_id","valid FK or blank",USER,"no","M-0001"),
 f("exclusion_type","What kind of restriction?","enum","allergy|religious|medical|disliked_ingredient|disliked_dish|never_used","one value",USER,"yes","allergy"),
 f("entity_type","Does this refer to an ingredient or dish?","enum","ingredient|dish|category","one value",USER,"yes","ingredient"),
 f("entity_id","What must be excluded?","varchar(40)","ingredient_master/dish_master ID","valid canonical ID",AI,"yes","ING_PEANUT","AI may match alias; user confirms"),
 f("severity","How strictly should it be avoided?","enum","preference|strict|medical_critical","one value",USER,"yes","medical_critical"),
 f("cross_contact_concern","Is cross-contact a concern?","boolean","true|false","required for allergy",USER,"no",True),
 f("user_note","Optional clarification","varchar(250)","free_text","max 250 chars",USER,"no","Avoid peanut oil too"),
 f("confirmed_by_user","Did the user confirm the normalized match?","boolean","true|false","must be true before hard filtering",HYBRID,"yes",True)],
"cooking_capability": [
 f("capability_id","Capability ID","uuid","UUID","unique",DERIVED,"yes","CC-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"),
 f("primary_cook_role","Who cooks most often?","enum","self|partner|family_member|domestic_cook|shared|none","one value",USER,"yes","shared"), f("cook_frequency","How often do you cook?","enum","never|rarely|1_2_days_week|3_5_days_week|daily|multiple_daily","one value",USER,"yes","daily"),
 f("skill_level","Cooking skill","enum","beginner|basic|intermediate|advanced","one value",USER,"yes","intermediate"), f("recipe_confidence","Confidence following recipes","tinyint","1-5","integer 1-5",USER,"yes",4), f("spice_handling_confidence","Confidence balancing Indian spices","tinyint","1-5","integer 1-5",USER,"no",4),
 f("preferred_complexity","Preferred recipe complexity","enum","very_simple|simple|moderate|elaborate","one value",USER,"yes","simple"), f("weekday_minutes","Weekday cooking time available","smallint","0-240","integer 0-240",USER,"yes",35), f("weekend_minutes","Weekend cooking time available","smallint","0-360","integer 0-360",USER,"yes",90),
 f("equipment_ids","Equipment available","json array","equipment_master.equipment_id[]","unique IDs",USER,"yes",'["EQ_GAS","EQ_PRESSURE_COOKER","EQ_MIXER"]'), f("novelty_willingness","Willingness to try new dishes","tinyint","1-5","integer 1-5",USER,"yes",4)],
"health_goals": [
 f("goal_id","Goal ID","uuid","UUID","unique",DERIVED,"yes","HG-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"), f("member_id","Whose goal?","uuid","members.member_id","valid FK",USER,"yes","M-0001"),
 f("goal_code","Primary nutrition goal","enum","weight_loss|weight_gain|maintenance|high_protein|diabetic_friendly|bp_friendly|high_fiber|low_carb|low_oil|low_sugar|child_nutrition|elder_nutrition|postpartum|recovery|fitness","one value per row",USER,"yes","high_protein"),
 f("priority","How important is this goal?","tinyint","1-5","integer 1-5",USER,"yes",4), f("protein_target_g","Daily protein target, if professionally set","decimal(5,1) nullable","grams/day","0-300; blank if unknown",USER,"no",75),
 f("clinician_advised","Was this advised by a qualified clinician?","boolean","true|false","boolean",USER,"no",False), f("active_from","Start date","date","YYYY-MM-DD","valid date",USER,"yes","2026-07-01"), f("active_to","End date, if temporary","date nullable","YYYY-MM-DD","after active_from or blank",USER,"no","")],
"meal_history": [
 f("meal_event_id","Meal event ID","uuid","UUID","unique",DERIVED,"yes","ME-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"), f("meal_date","When was it eaten?","date","last 30 days","date <= today; intake window 30 days",USER,"yes","2026-07-30"),
 f("meal_slot","Which meal?","enum","breakfast|lunch|dinner|snack","one value",USER,"yes","dinner"), f("dish_raw_name","What did you call the dish?","varchar(120)","free_text","1-120 chars",USER,"yes","Rajma chawal"), f("dish_alias_local","Local/household name","varchar(120)","free_text","max 120 chars",USER,"no","Rajmah rice"),
 f("canonical_dish_id","Normalized dish","varchar(40)","dish_master.dish_id","AI candidate + user validation",AI,"yes","DISH_RAJMA_CHWAL"), f("source_mode","How was it obtained?","enum","home_cooked|ordered|restaurant|packed_from_home|ready_to_eat","one value",USER,"yes","home_cooked"),
 f("portion_size","Approximate portion","enum","small|medium|large|shared|unknown","one value",USER,"no","medium"), f("frequency_count","How many times in the 30-day window?","smallint","1-30","integer 1-30; use one event/date where possible",USER,"yes",2),
 f("satisfaction_score","How satisfying was it?","tinyint","1-5","integer 1-5",USER,"no",5), f("sentiment","Overall reaction","enum","liked|neutral|disliked","one value",USER,"yes","liked"), f("repeat_desire","Would you want it again soon?","enum","yes|maybe|no|never_again","one value",USER,"yes","yes"),
 f("leftover_level","How much was left?","enum","none|some|a_lot|unknown","one value",USER,"no","some"), f("issue_flags","Any issues?","json array","too_spicy|too_oily|too_heavy|too_slow|too_expensive|none","unique values",USER,"no",'["none"]'),
 f("context_flags","Context","json array","weekday|weekend|festival|fasting|hot_weather|cold_weather|rainy","unique values",HYBRID,"no",'["weekday"]')],
"meal_consumers": [
 f("meal_consumer_id","Link ID","uuid","UUID","unique",DERIVED,"yes","MC-0001"), f("meal_event_id","Meal event","uuid","meal_history.meal_event_id","valid FK",DERIVED,"yes","ME-0001"), f("member_id","Who ate it?","uuid","members.member_id","valid FK",USER,"yes","M-0001"), f("member_sentiment","Their reaction","enum","liked|neutral|disliked|not_rated","one value",USER,"yes","liked"), f("portion_size","Their portion","enum","small|medium|large|unknown","one value",USER,"no","medium"), f("finished_portion","Did they finish?","enum","yes|partly|no|unknown","one value",USER,"no","yes")],
"dish_preferences": [
 f("dish_preference_id","Preference ID","uuid","UUID","unique",DERIVED,"yes","DP-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"), f("member_id","Member; blank for household","uuid nullable","members.member_id","valid FK or blank",USER,"no","M-0001"), f("dish_id","Dish","varchar(40)","dish_master.dish_id","valid canonical ID",AI,"yes","DISH_IDLI"),
 f("preference_type","How do you feel about it?","enum","favorite|liked|neutral|disliked|never_again|festive|comfort","one value",USER,"yes","favorite"), f("preference_score","Strength","tinyint","1-5","integer 1-5",USER,"yes",5), f("evidence_source","Where did this come from?","enum","onboarding|meal_history|rating|behavior_inference","one value",HYBRID,"yes","onboarding"), f("confidence_score","System confidence","decimal(3,2)","0-1","number 0-1",HYBRID,"yes",1.0), f("confirmed_by_user","User-confirmed?","boolean","true|false","boolean",HYBRID,"yes",True)],
"festival_seasonal": [
 f("occasion_pref_id","Record ID","uuid","UUID","unique",DERIVED,"yes","FS-0001"), f("household_id","Household","uuid","households.household_id","valid FK",DERIVED,"yes","HH-0001"), f("occasion_id","Festival/season/occasion","varchar(40)","occasion_master.occasion_id","valid FK",USER,"yes","FEST_DIWALI"), f("dish_ids","Foods associated with it","json array","dish_master.dish_id[]","max 20 unique IDs",AI,"yes",'["DISH_KHEER","DISH_MATHRI"]'), f("preference_strength","Importance","tinyint","1-5","integer 1-5",USER,"yes",5), f("restriction_codes","Special rules","json array","fasting|no_onion_garlic|satvik|vegetarian|none","unique values",USER,"no",'["vegetarian"]'), f("user_confirmed","Confirmed normalized dishes?","boolean","true|false","must be true to personalize",HYBRID,"yes",True)],
"recommendation_events": [
 f("event_id","Event ID","uuid","UUID","unique",USAGE,"yes","EV-0001"), f("user_id","Acting user","uuid","users.user_id","valid FK",USAGE,"yes","U-0001"), f("household_id","Household","uuid","households.household_id","valid FK",USAGE,"yes","HH-0001"), f("dish_id","Recommended dish","varchar(40)","dish_master.dish_id","valid FK",USAGE,"yes","DISH_POHA"),
 f("event_type","What happened?","enum","impression|opened|saved|planned|cooked|ordered|skipped|rejected|rated|locked|substituted","one value",USAGE,"yes","cooked"), f("event_at","When?","timestamp","ISO-8601 UTC","valid timestamp",USAGE,"yes","2026-08-01T07:40:00Z"), f("meal_slot","Recommendation slot","enum","breakfast|lunch|dinner|snack","one value",USAGE,"yes","breakfast"), f("context_json","Captured context","json","weekday/weather/season/festival","valid JSON; controlled keys",HYBRID,"no",'{"day_type":"weekday","weather":"rainy"}'), f("recommendation_rank","Rank shown","smallint","1-100","integer 1-100",USAGE,"no",1), f("reason_code","Why recommended","varchar(40)","reason_master.reason_code","valid FK",HYBRID,"no","QUICK_WEEKDAY"), f("feedback_score","Explicit rating","tinyint nullable","1-5","integer 1-5 or blank",USAGE,"no",5), f("substitute_dish_id","Replacement dish","varchar(40) nullable","dish_master.dish_id","valid FK or blank",USAGE,"no","")]
}

samples = {
"households": [
["HH-0001","Sharma Home","CITY_DELHI","ST_DL",'["ST_PB","ST_DL"]',"nuclear_family",4,2,1,1,0,"shared",18000,"2026-07-01T09:00:00Z"],
["HH-0002","Iyer Flat","CITY_BENGALURU","ST_KA",'["ST_TN"]',"couple",2,2,0,0,0,"shared",14000,"2026-07-02T10:00:00Z"],
["HH-0003","Banerjee Family","CITY_KOLKATA","ST_WB",'["ST_WB"]',"joint_family",6,3,1,2,0,"elder",24000,"2026-07-03T11:00:00Z"],
["HH-0004","Patel PG","CITY_AHMEDABAD","ST_GJ",'["ST_GJ"]',"hostel_pg",1,1,0,0,0,"self",9000,"2026-07-04T12:00:00Z"],
["HH-0005","Khan Household","CITY_HYDERABAD","ST_TG",'["ST_UP","ST_TG"]',"multigenerational",5,3,1,1,0,"shared",22000,"2026-07-05T13:00:00Z"]],
"users": [[f"U-000{i}",f"HH-000{i}",n,a,g,p,True,"mvp",True,f"2026-07-0{i}T09:05:00Z"] for i,(n,a,g,p) in enumerate([
 ("Asha","35_44","woman",'["working_professional","taste_first"]'),("Karthik","25_34","man",'["couple","health_focused"]'),("Moumita","45_59","woman",'["joint_family","tradition_first"]'),("Neel","18_24","man",'["student","budget_focused"]'),("Sana","25_34","woman",'["new_parent","convenience_first"]')],1)],
"members": [
["M-0001","HH-0001","U-0001","Asha","adult",'["breakfast","dinner"]',True,"Carries lunch to office"],
["M-0002","HH-0002","U-0002","Karthik","adult",'["breakfast","lunch","dinner"]',True,"Runs in the morning"],
["M-0003","HH-0003","U-0003","Moumita","adult",'["lunch","dinner"]',True,"Prefers seasonal fish"],
["M-0004","HH-0004","U-0004","Neel","adult",'["breakfast","dinner","snack"]',True,"Limited PG kitchen"],
["M-0005","HH-0005","U-0005","Sana","postpartum",'["breakfast","lunch","dinner"]',True,"Needs quick family meals"]],
"food_preferences": [
["FP-0001","HH-0001","M-0001","vegetarian",'["navratri"]',3,2,2,"balanced",6,1,1,'["DISH_RAJMA_CHWAL","DISH_KHICHDI"]'],
["FP-0002","HH-0002","M-0002","vegetarian",'["none"]',4,2,2,"mostly_gravy",6,1,0,'["DISH_CURD_RICE","DISH_DOSA"]'],
["FP-0003","HH-0003","M-0003","non_vegetarian",'["none"]',3,3,3,"mostly_gravy",7,1,0,'["DISH_MACHER_JHOL","DISH_KHICHDI"]'],
["FP-0004","HH-0004","M-0004","jain",'["weekly"]',2,3,2,"mostly_dry",3,2,2,'["DISH_DHOKLA","DISH_THEPLA"]'],
["FP-0005","HH-0005","M-0005","non_vegetarian",'["ramadan"]',4,3,3,"mostly_gravy",5,1,1,'["DISH_HALEEM","DISH_KHICHDI"]']],
"regional_taste": [
["RT-0001","HH-0001","REG_PUNJABI",5,"balanced","heritage","origin"], ["RT-0002","HH-0002","REG_TAMIL",5,"regional_purist","heritage","origin"], ["RT-0003","HH-0003","REG_BENGALI",5,"regional_purist","heritage","origin"], ["RT-0004","HH-0004","REG_GUJARATI",5,"balanced","heritage","origin"], ["RT-0005","HH-0005","REG_HYDERABADI",4,"flexible","regular","current_location"]],
"exclusions": [
["EX-0001","HH-0001","M-0001","allergy","ingredient","ING_PEANUT","medical_critical",True,"Avoid peanut oil too",True], ["EX-0002","HH-0002","M-0002","disliked_ingredient","ingredient","ING_BRINJAL","preference",False,"Texture dislike",True], ["EX-0003","HH-0003","M-0003","medical","category","CAT_HIGH_SODIUM","strict",False,"Clinician advised",True], ["EX-0004","HH-0004","M-0004","religious","ingredient","ING_ONION","strict",False,"Jain meals",True], ["EX-0005","HH-0005","M-0005","disliked_dish","dish","DISH_KARELA_SABZI","preference",False,"Too bitter",True]],
"cooking_capability": [
["CC-0001","HH-0001","shared","daily","intermediate",4,4,"simple",35,90,'["EQ_GAS","EQ_PRESSURE_COOKER","EQ_MIXER"]',4], ["CC-0002","HH-0002","self","daily","advanced",5,5,"moderate",45,120,'["EQ_INDUCTION","EQ_PRESSURE_COOKER","EQ_MIXER"]',4], ["CC-0003","HH-0003","domestic_cook","multiple_daily","advanced",4,5,"elaborate",60,150,'["EQ_GAS","EQ_PRESSURE_COOKER","EQ_OTG"]',3], ["CC-0004","HH-0004","self","3_5_days_week","beginner",3,2,"very_simple",20,45,'["EQ_INDUCTION","EQ_MICROWAVE"]',5], ["CC-0005","HH-0005","shared","daily","intermediate",4,4,"simple",30,75,'["EQ_GAS","EQ_PRESSURE_COOKER","EQ_AIR_FRYER"]',3]],
"health_goals": [
["HG-0001","HH-0001","M-0001","high_protein",4,75,False,"2026-07-01",""], ["HG-0002","HH-0002","M-0002","fitness",5,95,False,"2026-07-02",""], ["HG-0003","HH-0003","M-0003","bp_friendly",5,"",True,"2026-07-03",""], ["HG-0004","HH-0004","M-0004","maintenance",3,"",False,"2026-07-04",""], ["HG-0005","HH-0005","M-0005","postpartum",5,"",True,"2026-07-05","2026-12-31"]],
"meal_history": [
["ME-0001","HH-0001","2026-07-30","dinner","Rajma chawal","Rajmah rice","DISH_RAJMA_CHWAL","home_cooked","medium",2,5,"liked","yes","some",'["none"]','["weekday"]'],
["ME-0002","HH-0002","2026-07-29","breakfast","Masala dosa","Dose","DISH_MASALA_DOSA","home_cooked","large",1,4,"liked","yes","none",'["too_slow"]','["weekday"]'],
["ME-0003","HH-0003","2026-07-28","lunch","Macher jhol","Machher jhol","DISH_MACHER_JHOL","home_cooked","medium",3,5,"liked","yes","none",'["none"]','["weekend"]'],
["ME-0004","HH-0004","2026-07-27","snack","Dhokla","Khaman","DISH_DHOKLA","ordered","small",2,3,"neutral","maybe","some",'["too_expensive"]','["rainy"]'],
["ME-0005","HH-0005","2026-07-26","dinner","Chicken biryani","Hyderabadi biryani","DISH_CHICKEN_BIRYANI","home_cooked","large",1,5,"liked","yes","a_lot",'["too_heavy"]','["weekend"]']],
"meal_consumers": [[f"MC-000{i}",f"ME-000{i}",f"M-000{i}",s,p,fin] for i,(s,p,fin) in enumerate([("liked","medium","yes"),("liked","large","yes"),("liked","medium","yes"),("neutral","small","partly"),("liked","large","yes")],1)],
"dish_preferences": [[f"DP-000{i}",f"HH-000{i}",f"M-000{i}",d,t,sc,"onboarding",1.0,True] for i,(d,t,sc) in enumerate([("DISH_IDLI","favorite",5),("DISH_CURD_RICE","comfort",5),("DISH_MACHER_JHOL","favorite",5),("DISH_DHOKLA","liked",4),("DISH_HALEEM","festive",5)],1)],
"festival_seasonal": [[f"FS-000{i}",f"HH-000{i}",o,d,5,r,True] for i,(o,d,r) in enumerate([("FEST_DIWALI",'["DISH_KHEER","DISH_MATHRI"]','["vegetarian"]'),("FEST_PONGAL",'["DISH_VEN_PONGAL","DISH_SAKKARAI_PONGAL"]','["vegetarian"]'),("FEST_DURGA_PUJA",'["DISH_KHICHDI","DISH_LABRA"]','["vegetarian"]'),("SEASON_SUMMER",'["DISH_AAM_RAS","DISH_CHAAS"]','["none"]'),("FEST_RAMADAN",'["DISH_HALEEM","DISH_FRUIT_CHAAT"]','["fasting"]')],1)],
"recommendation_events": [[f"EV-000{i}",f"U-000{i}",f"HH-000{i}",d,e,f"2026-08-0{i}T07:40:00Z",m,c,rk,rs,fb,sub] for i,(d,e,m,c,rk,rs,fb,sub) in enumerate([
("DISH_POHA","cooked","breakfast",'{"day_type":"weekday","weather":"rainy"}',1,"QUICK_WEEKDAY",5,""), ("DISH_RAGI_DOSA","saved","breakfast",'{"day_type":"weekend","weather":"warm"}',2,"HIGH_PROTEIN",None,""), ("DISH_SHUKTO","planned","lunch",'{"day_type":"weekend","season":"monsoon"}',1,"REGIONAL_FIT",None,""), ("DISH_PANEER_WRAP","substituted","dinner",'{"day_type":"weekday"}',3,"QUICK_WEEKDAY",4,"DISH_JAIN_PANEER_WRAP"), ("DISH_KHICHDI","rated","dinner",'{"day_type":"weekday","mode":"recovery"}',1,"HEALTH_GOAL",5,"")],1)]
}

db_tables = {"households":"household_profile","users":"user_profile","members":"household_member","food_preferences":"food_preference","regional_taste":"regional_taste","exclusions":"food_exclusion","cooking_capability":"cooking_capability","health_goals":"nutrition_goal","meal_history":"meal_event","meal_consumers":"meal_event_member","dish_preferences":"dish_preference","festival_seasonal":"occasion_preference","recommendation_events":"recommendation_event"}

wb = Workbook(); wb.remove(wb.active)
navy="17365D"; blue="D9EAF7"; green="E2F0D9"; orange="FCE4D6"; gray="E7E6E6"; red="F4CCCC"; white="FFFFFF"
thin = Side(style="thin", color="C9D2DC")

def style_sheet(ws, freeze="A2"):
    ws.freeze_panes=freeze
    ws.auto_filter.ref=ws.dimensions
    for cell in ws[1]:
        cell.fill=PatternFill("solid", fgColor=navy); cell.font=Font(color=white,bold=True); cell.alignment=Alignment(wrap_text=True,vertical="center")
    ws.row_dimensions[1].height=32
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment=Alignment(vertical="top",wrap_text=True); c.border=Border(bottom=thin)
    for col in range(1,ws.max_column+1):
        vals=[str(ws.cell(r,col).value or "") for r in range(1,min(ws.max_row,80)+1)]
        ws.column_dimensions[get_column_letter(col)].width=min(48,max(11,max(map(len,vals))+2))

readme=wb.create_sheet("README")
readme.append(["section","guidance"])
for r in [
 ("Purpose","Database-ready Indian household meal-habits intake and ML-event template. Operational sheets contain five linked sample records."),
 ("Data strategy","Keep raw user wording, canonical IDs, source category, consent and event time. Normalize aliases without overwriting raw input."),
 ("How to use","Duplicate the workbook, delete sample rows only in blue DATA_* tabs, preserve headers and IDs, and load one table at a time."),
 ("Safety","Allergies and medical restrictions are user-confirmed hard constraints. AI may propose a match but must never invent one."),
 ("ML readiness","Use immutable event rows for exposures and actions; do not replace history with current preference snapshots.")]: readme.append(r)
style_sheet(readme)

for sheet, fields in schemas.items():
    ws=wb.create_sheet("DATA_"+sheet[:24])
    ws.append([x["field_name"] for x in fields])
    for row in samples[sheet]: ws.append(row)
    ws.sheet_properties.tabColor="5B9BD5"
    style_sheet(ws)
    # Simple native Excel validation for common scalar enums.
    for idx, spec in enumerate(fields,1):
        allowed=spec["allowed_values"]
        if "|" in allowed and len(allowed)<240 and not allowed.endswith("[]"):
            formula='"'+allowed.replace('|',',')+'"'
            dv=DataValidation(type="list",formula1=formula,allow_blank=spec["required_yes_no"]!="yes")
            dv.error="Choose a value from the list"; dv.errorTitle="Invalid value"; dv.promptTitle=spec["field_name"]; dv.prompt=spec["user_question"]
            ws.add_data_validation(dv); dv.add(f"{get_column_letter(idx)}2:{get_column_letter(idx)}1000")

dd=wb.create_sheet("FIELD_DICTIONARY")
dd.append(["sheet_name","field_name","user_question","db_table","db_column","data_type","allowed_values","validation_rule","source_category","required_yes_no","example_value","notes"])
for sheet, fields in schemas.items():
    for x in fields:
        dd.append(["DATA_"+sheet[:24],x["field_name"],x["user_question"],db_tables[sheet],x["field_name"],x["data_type"],x["allowed_values"],x["validation_rule"],x["source_category"],x["required_yes_no"],x["example_value"],x["notes"]])
style_sheet(dd)
dd.sheet_properties.tabColor="70AD47"

q=wb.create_sheet("QUESTIONNAIRE")
q.append(["order","stage","question","answer_type","maps_to","show_if","friction_note"])
questions=[
(1,"mvp","Where do you live now?","city_search","households.current_city_id","always","State is derived"),(2,"mvp","How many people usually eat at home, by age group?","four_counts","households.*_count","always","One compact step"),(3,"mvp","What best describes your household?","single_select","households.living_setup","always","Persona seed"),(4,"mvp","What is your household's main eating pattern?","single_select","food_preferences.diet_pattern","always","Ask member exceptions later"),(5,"mvp","Any allergies or foods that must never be suggested?","search_multi_select","exclusions","always","Require severity + confirmation"),
(6,"mvp","Which regional cuisines feel like home?","ranked_multi_select","regional_taste","always","Show Indian region cards"),(7,"mvp","Preferred spice and oil levels?","two_sliders","food_preferences.spice_level/oil_level","always","1-5 anchors"),(8,"mvp","Who cooks, and how much time is available on weekdays?","compact_select_numeric","cooking_capability","always","Combines capability fields"),(9,"mvp","What equipment can you use?","icon_multi_select","cooking_capability.equipment_ids","always","Preseed common Indian equipment"),(10,"mvp","Pick up to five dishes your household loves.","dish_search","dish_preferences","always","Alias-aware canonical search"),
(11,"full","Who eats each home meal regularly?","member_matrix","members.meal_participation","full path","Member-aware ranking"),(12,"full","How often are meals home-cooked, ordered, or eaten out?","three_counts","food_preferences.*_meals_week","full path","Use weekly counts"),(13,"full","Any religious, medical, or dislike-based exclusions?","structured_repeat","exclusions","full path","Separate hard rules from dislikes"),(14,"full","What health goals should meals support?","goal_cards","health_goals","full path","No medical inference"),(15,"full","What did you eat in the last 30 days?","repeat_event","meal_history + meal_consumers","full path","Offer calendar/import, retain date"),
(16,"full","Which festival or seasonal foods matter to your home?","occasion_dish_picker","festival_seasonal","full path","Region-aware suggestions"),(17,"full","How confident are you with recipes and spices?","two_sliders","cooking_capability","full path","1-5 anchors"),(18,"full","Which foods are comforting, festive, disliked, or never again?","categorized_dish_picker","dish_preferences","full path","One dish per preference row")]
for row in questions:q.append(row)
style_sheet(q); q.sheet_properties.tabColor="70AD47"

grp=wb.create_sheet("DB_GROUPING")
grp.append(["domain","tables","grain","primary_use","key_relationship"])
for row in [
("user_profile","user_profile","one row per app user","consent, demographics, persona","user_profile.household_id -> household_profile"),("household_profile","household_profile; household_member","one row per household/member","shared context and decision model","member.household_id"),("meal_history","meal_event; meal_event_member","one row per dish event/member consumption","sequence and household acceptance","meal_event_member.meal_event_id"),("dish_preferences","dish_preference","one row per household/member/dish signal","cold start and explicit taste","canonical dish_id"),("exclusions_allergies","food_exclusion","one row per constraint target","hard filters vs soft dislikes","member_id nullable"),
("regional_taste","regional_taste","one row per household/region","regional ranking and authenticity","cuisine_region_id"),("cooking_capability","cooking_capability","one row per household snapshot","feasibility filters","equipment IDs"),("nutrition_health_goals","nutrition_goal","one row per member/goal/time period","goal-aware ranking","active dates"),("contextual_behavior","recommendation_event","one immutable user action","time/context learning","user + dish + timestamp"),("persona","user_profile.persona_codes","multi-valued controlled codes","cold start segmentation","persona master"),("feedback_learning","recommendation_event; dish_preference","event log + current aggregate","ranking labels and feedback","derive snapshots from events")]: grp.append(row)
style_sheet(grp)

def list_sheet(name, headers, rows, color):
    ws=wb.create_sheet(name); ws.append(headers)
    for r in rows: ws.append(r)
    style_sheet(ws); ws.sheet_properties.tabColor=color; return ws

list_sheet("MVP_ONBOARDING",["order","field_or_group","why_required","destination"],[(1,"current_city_id","local availability and regional context","household_profile"),(2,"living_setup + member counts","household-aware portions","household_profile"),(3,"diet_pattern","base candidate filtering","food_preference"),(4,"allergies + strict exclusions","safety hard filters","food_exclusion"),(5,"regional cuisine preferences","cold-start taste","regional_taste"),(6,"spice_level + oil_level","fast sensory fit","food_preference"),(7,"primary cook + weekday minutes","recipe feasibility","cooking_capability"),(8,"equipment_ids","feasibility hard filters","cooking_capability"),(9,"3-5 loved dishes","strong positive seeds","dish_preference"),(10,"personalization consent","lawful product behavior","user_profile")],"FFC000")
list_sheet("FULL_ONBOARDING",["module","fields","priority","note"],[("identity_household","all household and member fields","required","Counts must reconcile"),("food_habits","diet, fasting, taste sliders, source mix","required","Member overrides allowed"),("regional_taste","ranked regions, authenticity, familiarity","required","Use seeded taxonomy"),("constraints","allergy, religious, medical, dislikes, never-used","required","Confirm normalized entity"),("cooking","ownership, frequency, skill, time, equipment, novelty","required","Feasibility layer"),("nutrition","member goals, dates, clinician-advised flag","optional","Do not infer diagnosis"),("meal_history","30-day dated dish events + consumers + reaction","recommended","Highest-value explicit history"),("occasion","festival, fasting and seasonal dishes","optional","Local relevance"),("dish_preferences","favorite/comfort/festive/disliked/never again","recommended","Canonical dish IDs"),("context","Do not ask; log later","later","See USAGE_LATER")],"FFC000")

masters=[
("dish_master","dish_id, canonical names, transliterations, course, veg class, ingredients, region","Before onboarding","Version taxonomy; never merge raw input destructively"),("dish_alias_master","alias, language, script, region, canonical dish_id, confidence","Before onboarding","Covers idli/इडली, dose/dosa, rajmah/rajma"),("ingredient_master","ingredient_id, names, aliases, allergen group, dietary flags","Before onboarding","Needed for exclusions and recipe composition"),("cuisine_region_master","national, state, subregional and community cuisines","Before onboarding","Avoid collapsing all South/North Indian food"),("location_master","state/UT, district, city, PIN mappings","Before onboarding","Derive state from city"),("equipment_master","pressure cooker, kadai, tawa, mixer, idli stand, OTG, etc.","Before onboarding","Capability matching"),("occasion_master","festivals, fasting periods, seasons, regional applicability","Before onboarding","Calendar dates may vary annually"),("persona_master","controlled persona codes and descriptions","Before onboarding","Multi-label, not mutually exclusive"),("nutrition_rule_master","goal-to-nutrient rules with qualified review","Before launch","Health claims require governance"),("reason_master","recommendation reason codes and display copy","Before launch","Supports explainability"),("unit_portion_master","Indian household units and portion conversions","Before history import","katori, roti, glass, ladle"),("weather_season_master","Indian seasons and weather buckets by geography","Before contextual ranking","External context only after consent")]
list_sheet("MASTER_SEED",["master","minimum_content","timing","governance_note"],masters,"A5A5A5")

ai_rows=[
("dish alias normalization","Suggest canonical_dish_id from raw/local name","User confirms ambiguous or safety-relevant match","Do not overwrite dish_raw_name"),("ingredient normalization","Map typed exclusion to ingredient/category ID","Always confirm hard exclusions","Never infer allergy severity"),("transliteration/language","Generate searchable variants","Keep original script and provenance","Do not treat spelling variant as new dish"),("dish enrichment","Suggest region, meal slot, ingredients and technique","Validate against curated master/recipe","No invented ingredient list"),("preference inference","Infer score from repeated/cooked/rated events","Use confidence + decay + user correction","Never convert weak behavior into hard exclusion"),("persona inference","Suggest convenience/taste/health orientation","Show/edit or keep as model feature only","No sensitive trait inference"),("portion/nutrition estimate","Estimate from dish and portion","Label estimate; use vetted composition tables","Not medical advice"),("conflict detection","Flag household preference conflicts","Resolve through member-level input","Do not silently prioritize one member"),("season/festival relevance","Suggest based on location/calendar","User can dismiss; use authoritative calendar","Do not infer religion from location/name")]
list_sheet("AI_RULES",["capability","ai_may_do","validation_required","never_do"],ai_rows,"A5A5A5")

usage_rows=[
("recommendation_impression","dish shown, rank, reason, model_version","recommendation_event","ranking exposure; never ask in onboarding"),("open_save_plan","opened/saved/planned","recommendation_event","intent signal"),("cook_order_complete","cooked/ordered/completed","recommendation_event","strong outcome signal"),("skip_reject_lock","skipped/rejected/never again","recommendation_event","distinguish transient skip from hard lock"),("rating_feedback","rating and optional reason","recommendation_event","explicit label"),("substitution","original and substitute dish/ingredient","recommendation_event","successful/failed swap learned from outcome"),("household_acceptance","which members ate/liked","meal_event_member","all vs some acceptance derived"),("timing_pattern","timestamp, meal slot, weekday/weekend","recommendation_event","derive pattern; do not ask"),("weather_context","coarse weather bucket","recommendation_event.context_json","consented external enrichment"),("season_festival_context","season/occasion active at event time","recommendation_event.context_json","hybrid seeded context"),("leftover_outcome","none/some/a lot after cooked meal","meal_event","prompt after meal only"),("cost_time_outcome","actual cost and cook time","meal_event extension","collect after action, not onboarding")]
list_sheet("USAGE_LATER",["signal","fields","destination","why_later"],usage_rows,"A5A5A5")

# Source legend and workbook-level polish.
legend=list_sheet("SOURCE_LEGEND",["source_category","definition","example"],[(USER,"Explicitly supplied by a person","diet_pattern"),(DERIVED,"Deterministically calculated from known fields","current_state_id from city"),(AI,"AI proposes normalization/enrichment; validation retained","canonical_dish_id"),(SEED,"Curated taxonomy/reference data loaded before use","dish_master"),(USAGE,"Captured from actual application interaction","recommendation skip"),(HYBRID,"Combines taxonomy, model and event evidence","preference confidence")],"A5A5A5")

for ws in wb.worksheets:
    ws.sheet_view.showGridLines=False
    ws.auto_filter.ref=ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.oddHeader.center.text = "&BIndian Meal Habits Data Template"
    ws.oddFooter.right.text = "Page &P of &N"

wb.save(OUT)

# Verification: reopen, check sheets and five sample rows in every DATA sheet.
check=load_workbook(OUT, data_only=False)
assert "FIELD_DICTIONARY" in check.sheetnames and "MVP_ONBOARDING" in check.sheetnames
for name in [n for n in check.sheetnames if n.startswith("DATA_")]:
    assert check[name].max_row == 6, (name, check[name].max_row)
print(OUT)
print(f"sheets={len(check.sheetnames)} data_sheets={len([n for n in check.sheetnames if n.startswith('DATA_')])} dictionary_fields={check['FIELD_DICTIONARY'].max_row-1}")
