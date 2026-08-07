"""Governed workbook-to-research ingestion for FooFoo synthetic training data.

The loader fails closed around a fixed private-target allowlist. Source-row retention is explicit,
rejected rows can be kept as compact lineage evidence, and production identities, plans, and
events are never written.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import tempfile
import uuid
from collections import Counter, defaultdict
from collections.abc import Iterable, Iterator
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TRANSFORMATION_VERSION = "foofoo-training-db-v1"
TARGETS = {
    "dish": "research.training_dishes",
    "household": "research.household_personas",
    "interaction": "research.interactions",
    "weekly_signal": "research.weekly_signals",
    "preference_edge": "research.household_preference_edges",
}
PRODUCTION_DENYLIST = {
    "auth.users",
    "public.profiles",
    "public.households",
    "public.household_members",
    "public.household_memberships",
    "public.week_plans",
    "public.plan_slots",
    "public.slates",
    "public.slate_items",
    "public.recommendation_events",
    "public.feedback_events",
    "public.outcome_events",
}
SOURCE_ROW_RETENTION = ("all", "rejected", "none")
REQUIRED_SHEETS = {
    "DATA_households",
    "DATA_users",
    "DATA_food_preferences",
    "DATA_meal_history",
    "DATA_recommendation_events",
}
CRITICAL_FIELDS: dict[str, tuple[str, ...]] = {
    "DATA_households": ("household_id", "current_state_id", "household_size"),
    "DATA_users": ("user_id", "household_id"),
    "DATA_members": ("member_id", "household_id"),
    "DATA_food_preferences": ("preference_id", "household_id"),
    "DATA_meal_history": (
        "meal_event_id",
        "household_id",
        "canonical_dish_id",
        "meal_date",
    ),
    "DATA_meal_consumers": ("meal_consumer_id", "meal_event_id", "member_id"),
    "DATA_recommendation_events": ("event_id", "user_id", "household_id", "dish_id"),
}


@dataclass
class SourceRow:
    """One physical workbook row and its validation outcome."""

    source_dataset: str
    source_file: str
    source_file_sha256: str
    sheet_name: str
    source_row_number: int
    source_record_key: str | None
    raw_payload: dict[str, Any]
    errors: list[str] = field(default_factory=list)

    @property
    def validation_status(self) -> str:
        """Return the persisted row status used to block normalization."""
        return "rejected" if self.errors else "accepted"

    @property
    def payload_sha256(self) -> str:
        """Return a deterministic hash of the source payload."""
        return _sha256_text(_canonical_json(self.raw_payload))


@dataclass(frozen=True)
class NormalizedRecord:
    """A private, versioned research record derived from checked-in training artifacts."""

    target_table: str
    record_key: str
    payload: dict[str, Any]
    confidence: float
    ontology_mapping_status: str
    source_lineage: tuple[dict[str, Any], ...]


def _json_value(value: Any) -> Any:
    """Convert spreadsheet values into stable JSON-compatible values."""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _canonical_json(value: Any) -> str:
    """Serialize a value deterministically for hashing and database storage."""
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)


def _sha256_bytes(value: bytes) -> str:
    """Return the lowercase SHA-256 digest for raw bytes."""
    return hashlib.sha256(value).hexdigest()


def _sha256_text(value: str) -> str:
    """Return the lowercase SHA-256 digest for UTF-8 text."""
    return _sha256_bytes(value.encode())


def _atomic_json(path: Path, value: dict[str, Any]) -> None:
    """Write a report atomically so interrupted runs never leave partial evidence."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w", encoding="utf-8", dir=path.parent, prefix=f".{path.name}.", delete=False
    ) as handle:
        json.dump(value, handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")
        temporary = handle.name
    os.replace(temporary, path)


def _file_descriptor(path: Path) -> dict[str, Any]:
    """Build a non-secret source descriptor with a content hash and byte count."""
    content = path.read_bytes()
    return {"name": path.name, "sha256": _sha256_bytes(content), "bytes": len(content)}


def read_workbook(path: Path, dataset: str) -> list[SourceRow]:
    """Read every non-empty physical workbook row with its real Excel row number."""
    descriptor = _file_descriptor(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    output: list[SourceRow] = []
    try:
        missing = REQUIRED_SHEETS - set(workbook.sheetnames)
        if missing:
            raise ValueError(f"{dataset} missing required sheets: {sorted(missing)}")
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            raw_headers = next(iterator, ())
            headers = [str(value or "").strip() for value in raw_headers]
            if not headers or any(not header for header in headers):
                raise ValueError(f"{dataset}/{worksheet.title} has blank headers")
            if len(headers) != len(set(headers)):
                raise ValueError(f"{dataset}/{worksheet.title} has duplicate headers")
            seen: set[str] = set()
            has_primary_key = worksheet.title.startswith("DATA_")
            for row_number, values in enumerate(iterator, start=2):
                if not any(value not in (None, "") for value in values):
                    continue
                payload = {
                    header: _json_value(value)
                    for header, value in zip(headers, values, strict=False)
                }
                key_value = payload.get(headers[0])
                key = None if key_value in (None, "") else str(key_value)
                errors: list[str] = []
                if has_primary_key:
                    if key is None:
                        errors.append("missing_primary_key")
                    elif key in seen:
                        errors.append("duplicate_primary_key")
                    else:
                        seen.add(key)
                for required in CRITICAL_FIELDS.get(worksheet.title, ()):
                    if payload.get(required) in (None, ""):
                        errors.append(f"missing_required:{required}")
                output.append(
                    SourceRow(
                        source_dataset=dataset,
                        source_file=path.name,
                        source_file_sha256=descriptor["sha256"],
                        sheet_name=worksheet.title,
                        source_row_number=row_number,
                        source_record_key=key,
                        raw_payload=payload,
                        errors=errors,
                    )
                )
    finally:
        workbook.close()
    return output


def validate_relationships(rows: list[SourceRow]) -> None:
    """Attach explicit orphan errors for workbook relationships used by the training transform."""
    grouped: dict[tuple[str, str], list[SourceRow]] = defaultdict(list)
    for row in rows:
        grouped[(row.source_dataset, row.sheet_name)].append(row)

    rules = (
        ("DATA_users", "household_id", "DATA_households", "household_id"),
        ("DATA_members", "household_id", "DATA_households", "household_id"),
        ("DATA_food_preferences", "household_id", "DATA_households", "household_id"),
        ("DATA_regional_taste", "household_id", "DATA_households", "household_id"),
        ("DATA_exclusions", "household_id", "DATA_households", "household_id"),
        ("DATA_cooking_capability", "household_id", "DATA_households", "household_id"),
        ("DATA_health_goals", "household_id", "DATA_households", "household_id"),
        ("DATA_meal_history", "household_id", "DATA_households", "household_id"),
        ("DATA_dish_preferences", "household_id", "DATA_households", "household_id"),
        ("DATA_festival_seasonal", "household_id", "DATA_households", "household_id"),
        ("DATA_recommendation_events", "household_id", "DATA_households", "household_id"),
        ("DATA_recommendation_events", "user_id", "DATA_users", "user_id"),
        ("DATA_meal_consumers", "meal_event_id", "DATA_meal_history", "meal_event_id"),
    )
    for dataset in {row.source_dataset for row in rows}:
        for child_sheet, child_key, parent_sheet, parent_key in rules:
            parents = {
                str(row.raw_payload[parent_key])
                for row in grouped.get((dataset, parent_sheet), [])
                if row.raw_payload.get(parent_key) not in (None, "")
            }
            for child in grouped.get((dataset, child_sheet), []):
                value = child.raw_payload.get(child_key)
                if value not in (None, "") and str(value) not in parents:
                    child.errors.append(f"orphan:{child_key}->{parent_sheet}.{parent_key}")


def verify_manifest(training_dir: Path) -> dict[str, Any]:
    """Validate every artifact checksum declared by the canonical training manifest."""
    manifest_path = training_dir / "manifest.json"
    manifest = json.loads(manifest_path.read_text())
    if manifest.get("synthetic_only") is not True:
        raise ValueError("training manifest must be synthetic_only=true")
    failures = []
    for name, expected in manifest.get("sha256", {}).items():
        path = training_dir / name
        actual = _sha256_bytes(path.read_bytes()) if path.is_file() else None
        if actual != expected:
            failures.append({"file": name, "expected": expected, "actual": actual})
    if failures:
        raise ValueError(f"training artifact checksum failures: {failures}")
    return manifest


def _jsonl(path: Path) -> Iterator[dict[str, Any]]:
    """Yield checked-in normalized artifact rows without loading a whole file at once."""
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if line.strip():
                try:
                    yield json.loads(line)
                except json.JSONDecodeError as exc:
                    raise ValueError(f"invalid JSON in {path}:{line_number}") from exc


def build_normalized_records(
    training_dir: Path,
    batch_id: str,
    source_files: list[dict[str, Any]],
) -> list[NormalizedRecord]:
    """Map domain-shaped artifacts to the fixed private research target allowlist."""
    forbidden = set(TARGETS.values()) & PRODUCTION_DENYLIST
    if forbidden or any(not target.startswith("research.") for target in TARGETS.values()):
        raise RuntimeError(f"unsafe training destination configuration: {sorted(forbidden)}")
    common = {
        "batch_id": batch_id,
        "transformation_version": TRANSFORMATION_VERSION,
        "source_files": source_files,
    }
    output: list[NormalizedRecord] = []
    ontology = json.loads((training_dir / "canonical_food_ontology.json").read_text())
    for value in ontology["dishes"]:
        output.append(
            NormalizedRecord(
                TARGETS["dish"],
                str(value["id"]),
                value,
                0.75,
                "mapped",
                (
                    {
                        **common,
                        "artifact": "canonical_food_ontology.json",
                        "datasets": value.get("source_datasets", []),
                    },
                ),
            )
        )
    artifact_specs = (
        ("household_features.jsonl", "household", "household_id", 0.70, "not_applicable"),
        ("interactions.jsonl", "interaction", "event_id", 0.70, "mapped"),
        ("weekly_signals.jsonl", "weekly_signal", "household_id", 0.65, "not_applicable"),
    )
    for filename, kind, key, confidence, mapping_status in artifact_specs:
        for value in _jsonl(training_dir / filename):
            output.append(
                NormalizedRecord(
                    TARGETS[kind],
                    str(value[key]),
                    value,
                    confidence,
                    mapping_status,
                    ({**common, "artifact": filename, "dataset": value.get("source_dataset")},),
                )
            )
    for index, value in enumerate(
        _jsonl(training_dir / "household_preference_graph.jsonl"), start=1
    ):
        key = _sha256_text(_canonical_json(value))
        output.append(
            NormalizedRecord(
                TARGETS["preference_edge"],
                key,
                value,
                0.65,
                "mapped" if str(value.get("source", "")).startswith("DISH_") else "not_applicable",
                (
                    {
                        **common,
                        "artifact": "household_preference_graph.jsonl",
                        "artifact_row": index,
                    },
                ),
            )
        )
    return output


def _bundle_identity(
    source_files: list[dict[str, Any]], manifest: dict[str, Any]
) -> tuple[str, str]:
    """Return the content-derived bundle hash and stable UUID batch identity."""
    digest = _sha256_text(
        _canonical_json(
            {
                "source_files": source_files,
                "manifest": manifest.get("sha256", {}),
                "transformation_version": TRANSFORMATION_VERSION,
            }
        )
    )
    return digest, str(uuid.uuid5(uuid.NAMESPACE_URL, f"foofoo:training:{digest}"))


def _blocked_summary(rows: Iterable[SourceRow]) -> dict[str, Any]:
    """Build a complete, compact rejected-row report with explicit reasons."""
    blocked = [row for row in rows if row.errors]
    return {
        "count": len(blocked),
        "by_reason": dict(
            sorted(Counter(error for row in blocked for error in row.errors).items())
        ),
        "rows": [
            {
                "dataset": row.source_dataset,
                "sheet": row.sheet_name,
                "row": row.source_row_number,
                "record_key": row.source_record_key,
                "reasons": sorted(set(row.errors)),
            }
            for row in blocked
        ],
    }


def deduplicate_normalized_records(
    records: Iterable[NormalizedRecord],
) -> tuple[list[NormalizedRecord], dict[str, int]]:
    """Collapse exact duplicate entities and fail on conflicting payloads for one natural key."""
    unique: dict[tuple[str, str], NormalizedRecord] = {}
    duplicates: Counter[str] = Counter()
    for record in records:
        key = (record.target_table, record.record_key)
        prior = unique.get(key)
        if prior is None:
            unique[key] = record
            continue
        if _canonical_json(prior.payload) != _canonical_json(record.payload):
            raise ValueError(
                f"conflicting normalized records for {record.target_table}/{record.record_key}"
            )
        duplicates[record.target_table] += 1
    return list(unique.values()), dict(sorted(duplicates.items()))


def build_ingestion(
    dataset_1: Path, dataset_2: Path, training_dir: Path
) -> tuple[dict[str, Any], list[SourceRow], list[NormalizedRecord]]:
    """Audit sources, verify derived artifacts, and build the deterministic load package."""
    manifest = verify_manifest(training_dir)
    paths = [("dataset_1", dataset_1), ("dataset_2", dataset_2)]
    source_files = [{"dataset": label, **_file_descriptor(path)} for label, path in paths]
    rows: list[SourceRow] = []
    for label, path in paths:
        rows.extend(read_workbook(path, label))
    validate_relationships(rows)
    bundle_hash, batch_id = _bundle_identity(source_files, manifest)
    generated_records = build_normalized_records(training_dir, batch_id, source_files)
    records, normalized_duplicates = deduplicate_normalized_records(generated_records)
    accepted = sum(row.validation_status == "accepted" for row in rows)
    report = {
        "batch_id": batch_id,
        "import_key": f"{manifest['version']}:{TRANSFORMATION_VERSION}:{bundle_hash[:16]}",
        "source_bundle_sha256": bundle_hash,
        "source_dataset_version": manifest["version"],
        "generation_version": manifest.get("generated_at", "unknown"),
        "transformation_version": TRANSFORMATION_VERSION,
        "synthetic_only": True,
        "source_files": source_files,
        "source_rows": {"total": len(rows), "accepted": accepted, "rejected": len(rows) - accepted},
        "normalized_records": {
            "total": len(records),
            "by_target": dict(sorted(Counter(record.target_table for record in records).items())),
            "exact_duplicates_skipped": sum(normalized_duplicates.values()),
            "duplicates_by_target": normalized_duplicates,
        },
        "blocked_rows": _blocked_summary(rows),
        "production_targets": [],
        "policy": {
            "production_denylist_enforced": True,
            "active_model_promotion_allowed": False,
            "destination": "private research/ml staging only",
        },
    }
    return report, rows, records


def _chunks(values: list[Any], size: int = 1000) -> Iterator[list[Any]]:
    """Yield bounded chunks for predictable database memory and statement sizes."""
    for start in range(0, len(values), size):
        yield values[start : start + size]


def retained_source_rows(
    source_rows: list[SourceRow], retention: str
) -> list[SourceRow]:
    """Select raw lineage rows for the requested storage profile.

    The normalized training records are unaffected. Keeping only rejected rows preserves
    actionable validation evidence without duplicating every accepted workbook payload in
    PostgreSQL.
    """
    if retention not in SOURCE_ROW_RETENTION:
        raise ValueError(f"unsupported source-row retention: {retention}")
    if retention == "all":
        return source_rows
    if retention == "rejected":
        return [row for row in source_rows if row.validation_status == "rejected"]
    return []


def load_postgres(
    connection: Any,
    package: dict[str, Any],
    source_rows: list[SourceRow],
    records: list[NormalizedRecord],
    *,
    source_row_retention: str = "all",
) -> dict[str, Any]:
    """Load one governed batch transactionally and return exact insert/update/skip counts."""
    from psycopg2.extras import execute_values

    batch_id = package["batch_id"]
    with connection.cursor() as cursor:
        cursor.execute("SELECT pg_advisory_xact_lock(hashtextextended(%s,0))", (batch_id,))
        cursor.execute(
            "SELECT to_regclass('research.training_source_rows'), "
            "to_regclass('ml.training_import_batches')"
        )
        existing_tables = cursor.fetchone()
        if not existing_tables or any(value is None for value in existing_tables):
            raise RuntimeError("migration 088 is not applied; refusing training load")
        cursor.execute(
            """INSERT INTO ml.training_import_batches(
                 id,import_key,source_bundle_sha256,source_dataset_version,generation_version,
                 transformation_version,source_files,source_row_count,accepted_source_row_count,
                 rejected_source_row_count,normalized_record_count)
               VALUES(%s,%s,%s,%s,%s,%s,%s::jsonb,%s,%s,%s,%s)
               ON CONFLICT(id) DO UPDATE SET status='loading',started_at=now(),completed_at=NULL""",
            (
                batch_id,
                package["import_key"],
                package["source_bundle_sha256"],
                package["source_dataset_version"],
                package["generation_version"],
                package["transformation_version"],
                _canonical_json(package["source_files"]),
                package["source_rows"]["total"],
                package["source_rows"]["accepted"],
                package["source_rows"]["rejected"],
                package["normalized_records"]["total"],
            ),
        )
        retained_rows = retained_source_rows(source_rows, source_row_retention)
        raw_values = [
            (
                batch_id,
                row.source_dataset,
                row.source_file,
                row.source_file_sha256,
                row.sheet_name,
                row.source_row_number,
                row.source_record_key,
                _canonical_json(row.raw_payload),
                row.payload_sha256,
                row.validation_status,
                _canonical_json(sorted(set(row.errors))),
            )
            for row in retained_rows
        ]
        for chunk in _chunks(raw_values):
            execute_values(
                cursor,
                """INSERT INTO research.training_source_rows(
                     batch_id,source_dataset,source_file,source_file_sha256,sheet_name,
                     source_row_number,source_record_key,raw_payload,raw_payload_sha256,
                     validation_status,validation_errors) VALUES %s
                   ON CONFLICT(batch_id,source_dataset,sheet_name,source_row_number) DO NOTHING""",
                chunk,
                template="(%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s,%s,%s::jsonb)",
            )

        cursor.execute(
            """SELECT target_table,record_key,payload_sha256,confidence
               FROM research.auto_training_records WHERE target_table=ANY(%s)""",
            (list(TARGETS.values()),),
        )
        existing = {
            (str(row[0]), str(row[1])): (str(row[2]), float(row[3]))
            for row in cursor.fetchall()
        }
        counts: dict[str, Counter[str]] = defaultdict(Counter)
        inserts: list[tuple[Any, ...]] = []
        updates: list[tuple[Any, ...]] = []
        for record in records:
            digest = _sha256_text(_canonical_json(record.payload))
            prior = existing.get((record.target_table, record.record_key))
            values = (
                record.target_table,
                record.record_key,
                _canonical_json(record.payload),
                digest,
                record.confidence,
                record.ontology_mapping_status,
                _canonical_json(list(record.source_lineage)),
            )
            if prior is None:
                inserts.append(
                    (
                        record.target_table,
                        record.record_key,
                        values[2],
                        digest,
                        record.confidence,
                        record.confidence,
                        record.confidence,
                        record.ontology_mapping_status,
                        batch_id,
                        batch_id,
                        package["source_dataset_version"],
                        package["generation_version"],
                        package["transformation_version"],
                        values[6],
                    )
                )
                counts[record.target_table]["inserted"] += 1
            elif prior[0] == digest or record.confidence < prior[1]:
                counts[record.target_table]["skipped"] += 1
            else:
                updates.append(values)
                counts[record.target_table]["updated"] += 1
        for chunk in _chunks(inserts):
            execute_values(
                cursor,
                """INSERT INTO research.auto_training_records(
                     target_table,record_key,payload,payload_sha256,source_type,generation_method,
                     confidence,confidence_band,ontology_mapping_status,ontology_version,
                     provenance_tags,first_batch_id,last_batch_id,synthetic_only,
                     source_dataset_version,generation_version,transformation_version,source_lineage)
                   VALUES %s""",
                chunk,
                template=(
                    "(%s,%s,%s::jsonb,%s,'expert_research_synthetic','foofoo-training-db-v1',%s,"
                    "CASE WHEN %s>=0.85 THEN 'high' WHEN %s>=0.65 THEN 'medium' ELSE 'low' END,"
                    "%s,'foofoo-training-v1',ARRAY['synthetic','workbook_import'],%s,%s,true,%s,%s,%s,%s::jsonb)"
                ),
                fetch=False,
                page_size=1000,
            )
        # Updates use executemany because each row carries its own WHERE key and confidence guard.
        cursor.executemany(
            """UPDATE research.auto_training_records SET payload=%s::jsonb,payload_sha256=%s,
                 confidence=%s,confidence_band=CASE WHEN %s>=0.85 THEN 'high'
                   WHEN %s>=0.65 THEN 'medium' ELSE 'low' END,
                 ontology_mapping_status=%s,last_batch_id=%s,version=version+1,updated_at=now(),
                 synthetic_only=true,source_dataset_version=%s,generation_version=%s,
                 transformation_version=%s,source_lineage=%s::jsonb
               WHERE target_table=%s AND record_key=%s AND confidence<=%s""",
            [
                (
                    payload,
                    digest,
                    confidence,
                    confidence,
                    confidence,
                    mapping,
                    batch_id,
                    package["source_dataset_version"],
                    package["generation_version"],
                    package["transformation_version"],
                    lineage,
                    target,
                    key,
                    confidence,
                )
                for target, key, payload, digest, confidence, mapping, lineage in updates
            ],
        )
        normalized_summary = {
            target: dict(sorted(counter.items()))
            for target, counter in sorted(counts.items())
        }
        load_summary = {
            "normalized": normalized_summary,
            "source_row_retention": source_row_retention,
            "retained_source_rows": len(retained_rows),
        }
        status = "completed_with_rejections" if package["source_rows"]["rejected"] else "completed"
        cursor.execute(
            """UPDATE ml.training_import_batches SET status=%s,load_summary=%s::jsonb,
                 completed_at=now() WHERE id=%s""",
            (status, _canonical_json(load_summary), batch_id),
        )
        cursor.execute(
            """SELECT validation_status,count(*) FROM research.training_source_rows
               WHERE batch_id=%s GROUP BY validation_status""",
            (batch_id,),
        )
        raw_db_counts = {str(row[0]): int(row[1]) for row in cursor.fetchall()}
    return {
        "status": status,
        "normalized": normalized_summary,
        "source_row_retention": source_row_retention,
        "raw_source_rows": raw_db_counts,
    }


def connect(dsn: str) -> Any:
    """Open the privileged PostgreSQL connection required by execute mode."""
    import psycopg2

    return psycopg2.connect(
        dsn, connect_timeout=15, application_name="foofoo-training-ingestion"
    )


def database_url() -> str:
    """Return a protected DB URL without ever printing its value."""
    for name in ("DATABASE_URL", "SUPABASE_DB_URL", "FOOFOO_SUPABASE_URI"):
        if value := os.getenv(name):
            return value
    raise RuntimeError(
        "execute mode requires DATABASE_URL, SUPABASE_DB_URL, or FOOFOO_SUPABASE_URI"
    )


def main(argv: list[str] | None = None) -> int:
    """Run audit/dry-run locally or execute an approved private-schema load."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--mode", choices=("audit", "dry_run", "execute"), default="audit")
    parser.add_argument("--dataset-1", type=Path, required=True)
    parser.add_argument("--dataset-2", type=Path, required=True)
    parser.add_argument("--training-dir", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument(
        "--source-row-retention",
        choices=SOURCE_ROW_RETENTION,
        default="all",
        help=(
            "Persist all, rejected-only, or no raw workbook rows; normalized records "
            "are unchanged."
        ),
    )
    args = parser.parse_args(argv)
    report, rows, records = build_ingestion(args.dataset_1, args.dataset_2, args.training_dir)
    report["mode"] = args.mode
    report["load"] = {"executed": False}
    if args.mode == "execute":
        connection = connect(database_url())
        try:
            report["load"] = {
                "executed": True,
                **load_postgres(
                    connection,
                    report,
                    rows,
                    records,
                    source_row_retention=args.source_row_retention,
                ),
            }
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
    elif args.mode == "dry_run":
        retained_rows = retained_source_rows(rows, args.source_row_retention)
        report["load"]["would_write"] = {
            "raw_source_rows": len(retained_rows),
            "source_row_retention": args.source_row_retention,
            "normalized_records": len(records),
            "production_records": 0,
        }
    _atomic_json(args.report, report)
    print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
