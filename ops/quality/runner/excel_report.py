"""Build validated Excel and ZIP evidence from FooFoo quality-test outputs.

The exporter is deliberately a reader: it never invents a user, journey event, recommendation,
or test result. Persona journeys are read from ``personas/<id>`` while the general quality gate is
read from ``test_results.json``. The source persona workbook is used only for traceability and for
copying the matching precomputed persona/weekly-plan rows into the result workbook.

Usage::

    python ops/quality/runner/excel_report.py REPORT_DIR --kind persona-journey
    python ops/quality/runner/excel_report.py REPORT_DIR --kind quality-gate
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import subprocess
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_SOURCE_WORKBOOK = REPO_ROOT / "data" / "source" / "Indian_Meal_Cohort_Persona_DB_v3.xlsx"
REQUIRED_SHEETS = (
    "Run_Summary",
    "Users",
    "Journey_Events",
    "Recommendations",
    "Final_Meal_Plans",
    "Test_Results",
    "Errors",
    "Source_Traceability",
    "Source_Personas",
    "Source_Meal_Plans",
)

LOGGER = logging.getLogger("foofoo.quality.excel_report")


@dataclass(frozen=True)
class ReportPaths:
    """Files produced for one report directory and uploaded by CI."""

    workbook: Path
    archive: Path
    manifest: Path
    summary: Path
    errors: Path


def _json_log(event: str, **fields: Any) -> None:
    """Emit one structured JSON log line without journey answers, credentials, or tokens."""
    LOGGER.info(json.dumps({"event": event, **fields}, default=str, sort_keys=True))


def _read_json(path: Path, default: Any = None) -> Any:
    """Read a JSON artifact, returning ``default`` only when the file does not exist."""
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def _git_head() -> str:
    """Return the checked-out short commit id for report traceability."""
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"], cwd=REPO_ROOT, text=True
        ).strip()
    except (OSError, subprocess.SubprocessError):
        return "unknown"


def _sha256(path: Path) -> str:
    """Return a file's SHA-256 digest without loading the whole file into memory."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_excel(value: Any) -> Any:
    """Convert nested values to readable cells and neutralize spreadsheet formulas."""
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _add_sheet(
    workbook: Workbook, title: str, headers: list[str], rows: Iterable[Iterable[Any]]
) -> None:
    """Create a styled, filterable worksheet from headers and row values."""
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([_safe_excel(value) for value in row])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        letter = column[0].column_letter
        width = max((len(str(cell.value or "")) for cell in column), default=8)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 12), 60)
        for cell in column[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if worksheet.max_row > 1 and worksheet.max_column > 0:
        table_name = "T" + "".join(ch for ch in title if ch.isalnum())[:240]
        table = Table(displayName=table_name, ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)


def _source_persona_id(summary: dict[str, Any]) -> str | None:
    """Resolve a source persona id from explicit metadata or the legacy ``real_pNN`` key."""
    explicit = summary.get("source_persona_id")
    if explicit:
        return str(explicit).upper()
    key = str(summary.get("key", ""))
    if key.startswith("real_p") and key[6:].isdigit():
        return f"P{int(key[6:]):02d}"
    return None


def _recommendation_rows(
    run_id: str, user_id: str, event: dict[str, Any], fallback_stage: str
) -> tuple[list[list[Any]], list[list[Any]]]:
    """Flatten one recommendation response into dish rows and final-plan plate rows."""
    timestamp = event.get("timestamp_utc")
    stage = event.get("stage_label") or fallback_stage
    response = event.get("response") or event
    http_status = response.get("status")
    body = response.get("body") if isinstance(response, dict) else None
    body = body if isinstance(body, dict) else {}
    request_id = body.get("request_id")
    recommendation_rows: list[list[Any]] = []
    plan_rows: list[list[Any]] = []
    for rank, plate in enumerate(body.get("plates") or [], start=1):
        if not isinstance(plate, dict):
            continue
        dish_ids = plate.get("hero_dish_ids") or []
        dish_names = plate.get("hero_dish_names") or []
        plan_rows.append(
            [
                run_id,
                user_id,
                timestamp,
                stage,
                request_id,
                rank,
                plate.get("plate_id"),
                plate.get("form"),
                " | ".join(map(str, dish_ids)),
                " | ".join(map(str, dish_names)),
                plate.get("support"),
                plate.get("plate_score"),
                body.get("warnings") or [],
            ]
        )
        width = max(len(dish_ids), len(dish_names), 1)
        for dish_rank in range(width):
            recommendation_rows.append(
                [
                    run_id,
                    user_id,
                    timestamp,
                    stage,
                    request_id,
                    http_status,
                    rank,
                    plate.get("plate_id"),
                    dish_rank + 1,
                    dish_ids[dish_rank] if dish_rank < len(dish_ids) else None,
                    dish_names[dish_rank] if dish_rank < len(dish_names) else None,
                    plate.get("form"),
                    plate.get("support"),
                    plate.get("plate_score"),
                ]
            )
    return recommendation_rows, plan_rows


def _load_persona_evidence(report_dir: Path, run_id: str) -> dict[str, list[list[Any]]]:
    """Read every completed or failed persona directory into workbook-ready rows."""
    users: list[list[Any]] = []
    journeys: list[list[Any]] = []
    recommendations: list[list[Any]] = []
    plans: list[list[Any]] = []
    tests: list[list[Any]] = []
    errors: list[list[Any]] = []
    personas_root = report_dir / "personas"
    if not personas_root.exists():
        personas_root = report_dir / "personas-ui" / "personas"
    if not personas_root.exists():
        return {
            "users": users,
            "journeys": journeys,
            "recommendations": recommendations,
            "plans": plans,
            "tests": tests,
            "errors": errors,
        }

    for persona_dir in sorted(path for path in personas_root.iterdir() if path.is_dir()):
        summary = _read_json(persona_dir / "summary.json", {}) or {}
        if not summary:
            errors.append(
                [
                    run_id,
                    None,
                    datetime.now(UTC).isoformat(),
                    "evidence",
                    "missing summary.json",
                    str(persona_dir),
                ]
            )
            continue
        user_id = str(summary.get("test_user_id") or summary.get("key") or persona_dir.name)
        user_type = summary.get("user_type") or "synthetic"
        started = summary.get("started_at_utc")
        completed = summary.get("completed_at_utc")
        source_id = _source_persona_id(summary)
        users.append(
            [
                run_id,
                user_id,
                user_type,
                summary.get("key"),
                summary.get("label"),
                source_id,
                started,
                completed,
                "pass" if summary.get("ok") else "fail",
            ]
        )

        steps = summary.get("steps") or []
        for sequence, step in enumerate(steps, start=1):
            journeys.append(
                [
                    run_id,
                    user_id,
                    user_type,
                    summary.get("key"),
                    started,
                    step.get("timestamp_utc") or started,
                    sequence,
                    step.get("label"),
                    step.get("action") or step.get("label"),
                    step.get("screenshot"),
                    "pass" if summary.get("ok") else "incomplete",
                ]
            )

        rec_events = _read_json(persona_dir / "recommendation_events.json", None)
        if not isinstance(rec_events, list):
            final_rec = _read_json(persona_dir / "recommendations.json", None)
            rec_events = (
                [
                    {
                        "timestamp_utc": completed,
                        "stage_label": "final-recommendations",
                        "response": final_rec,
                    }
                ]
                if final_rec is not None
                else []
            )
        for event in rec_events:
            rec_rows, plan_rows = _recommendation_rows(
                run_id, user_id, event, "final-recommendations"
            )
            recommendations.extend(rec_rows)
            plans.extend(plan_rows)

        actual_status = summary.get("recommendations_status")
        expected_status = summary.get("expect_status")
        status_matches = expected_status is None or actual_status == expected_status
        test_status = "pass" if summary.get("ok") and status_matches else "fail"
        tests.append(
            [
                run_id,
                "persona-journey",
                user_id,
                summary.get("key"),
                "complete journey and receive expected recommendation status",
                expected_status,
                actual_status,
                test_status,
                summary.get("error"),
                completed,
            ]
        )
        if summary.get("error"):
            errors.append(
                [
                    run_id,
                    user_id,
                    completed,
                    "persona-journey",
                    summary.get("error"),
                    str(persona_dir),
                ]
            )

    return {
        "users": users,
        "journeys": journeys,
        "recommendations": recommendations,
        "plans": plans,
        "tests": tests,
        "errors": errors,
    }


def _load_quality_evidence(
    report_dir: Path, run_id: str
) -> tuple[list[list[Any]], list[list[Any]]]:
    """Flatten the generic quality gate's steps and JUnit cases into test/error rows."""
    payload = _read_json(report_dir / "test_results.json", {}) or {}
    generated = (payload.get("meta") or {}).get("generated_at")
    tests: list[list[Any]] = []
    errors: list[list[Any]] = []
    for step in payload.get("steps") or []:
        cases = step.get("cases") or []
        if cases:
            for case in cases:
                tests.append(
                    [
                        run_id,
                        step.get("name"),
                        None,
                        None,
                        case.get("test"),
                        None,
                        None,
                        case.get("status"),
                        None,
                        generated,
                    ]
                )
        else:
            tests.append(
                [
                    run_id,
                    step.get("name"),
                    None,
                    None,
                    step.get("summary") or step.get("reason"),
                    None,
                    None,
                    step.get("status"),
                    step.get("reason"),
                    generated,
                ]
            )
        for failure in step.get("failures") or []:
            errors.append(
                [
                    run_id,
                    None,
                    generated,
                    step.get("name"),
                    failure.get("message"),
                    failure.get("test"),
                ]
            )
    return tests, errors


def _copy_source_rows(
    source_workbook: Path, persona_ids: set[str]
) -> tuple[list[str], list[list[Any]], list[str], list[list[Any]]]:
    """Copy matching persona and precomputed weekly-plan rows from the canonical source workbook."""
    source = load_workbook(source_workbook, read_only=True, data_only=True)

    def selected(sheet_name: str) -> tuple[list[str], list[list[Any]]]:
        """Return the header and rows whose persona_id is in the report's mapped persona set."""
        worksheet = source[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(iterator)]
        persona_index = headers.index("persona_id")
        if not persona_ids:
            return headers, []
        rows = [list(row) for row in iterator if str(row[persona_index]) in persona_ids]
        return headers, rows

    persona_headers, persona_rows = selected("Persona_Master_v3")
    plan_headers, plan_rows = selected("Weekly_Class_Plan_v3")
    source.close()
    return persona_headers, persona_rows, plan_headers, plan_rows


def _write_manifest(
    report_dir: Path,
    paths: ReportPaths,
    run_id: str,
    kind: str,
    environment: str,
    source_workbook: Path,
    generated_at: str,
) -> dict[str, Any]:
    """Write machine-readable provenance and a compact human-readable run summary."""
    source_entry = {
        "path": str(source_workbook.relative_to(REPO_ROOT))
        if source_workbook.is_relative_to(REPO_ROOT)
        else str(source_workbook),
        "sha256": _sha256(source_workbook) if source_workbook.exists() else None,
    }
    manifest = {
        "schema_version": "1.0",
        "run_id": run_id,
        "kind": kind,
        "environment": environment,
        "generated_at_utc": generated_at,
        "git_head": _git_head(),
        "source_workbook": source_entry,
        "workbook": paths.workbook.name,
        "archive": paths.archive.name,
    }
    paths.manifest.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    paths.summary.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest


def _validate_workbook(path: Path, *, require_users: bool) -> None:
    """Fail report publication when required sheets or persona evidence are absent."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(f"workbook is missing required sheets: {', '.join(missing)}")
    if require_users and workbook["Users"].max_row <= 1:
        raise ValueError("persona report contains no user rows")
    if require_users and workbook["Journey_Events"].max_row <= 1:
        raise ValueError("persona report contains no journey event rows")
    workbook.close()


def _write_archive(report_dir: Path, paths: ReportPaths, kind: str) -> None:
    """Create one directly downloadable ZIP, placing persona UI evidence under ui-artifacts/."""
    excluded = {paths.archive.resolve()}
    with zipfile.ZipFile(
        paths.archive, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6
    ) as archive:
        for item in sorted(report_dir.rglob("*")):
            if not item.is_file() or item.resolve() in excluded:
                continue
            relative = item.relative_to(report_dir)
            if kind == "persona-journey" and relative.parts[0] in {"personas", "report"}:
                arcname = Path("ui-artifacts") / relative
            else:
                arcname = relative
            archive.write(item, arcname)


def build_report(
    report_dir: Path,
    *,
    kind: str,
    source_workbook: Path = DEFAULT_SOURCE_WORKBOOK,
    run_id: str | None = None,
    environment: str = "local",
) -> ReportPaths:
    """Build, validate, and package a quality report from real artifacts already on disk."""
    report_dir = report_dir.resolve()
    report_dir.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now(UTC).isoformat()
    run_id = run_id or os.environ.get("GITHUB_RUN_ID") or report_dir.name
    safe_run = "".join(ch if ch.isalnum() or ch in "-_" else "-" for ch in run_id)
    prefix = "persona-journey" if kind == "persona-journey" else "quality-report"
    workbook_path = report_dir / "test-results.xlsx"
    archive_path = report_dir / f"{prefix}_{environment}_{date.today().isoformat()}_{safe_run}.zip"
    paths = ReportPaths(
        workbook=workbook_path,
        archive=archive_path,
        manifest=report_dir / "source-manifest.json",
        summary=report_dir / "run-summary.json",
        errors=report_dir / "errors.log",
    )
    paths.errors.touch(exist_ok=True)

    if not source_workbook.exists():
        raise FileNotFoundError(f"canonical source workbook is missing: {source_workbook}")

    persona = _load_persona_evidence(report_dir, run_id)
    quality_tests, quality_errors = _load_quality_evidence(report_dir, run_id)
    persona["tests"].extend(quality_tests)
    persona["errors"].extend(quality_errors)
    source_ids = {str(row[5]) for row in persona["users"] if row[5]}
    source_persona_headers, source_persona_rows, source_plan_headers, source_plan_rows = (
        _copy_source_rows(source_workbook, source_ids)
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary_payload = _read_json(report_dir / "persona_journeys_result.json", None)
    if summary_payload is None:
        summary_payload = _read_json(
            report_dir / "personas-ui" / "persona_journeys_result.json", {}
        )
    summary_payload = summary_payload or {}
    quality_payload = _read_json(report_dir / "test_results.json", {}) or {}
    verdict = quality_payload.get("verdict") or {}
    summary_rows = [
        ["run_id", run_id],
        ["report_kind", kind],
        ["generated_at_utc", generated_at],
        ["environment", environment],
        ["git_head", _git_head()],
        ["journey_status", summary_payload.get("status")],
        ["users_total", len(persona["users"])],
        ["users_passed", sum(1 for row in persona["users"] if row[-1] == "pass")],
        ["tests_total", len(persona["tests"])],
        ["quality_score", verdict.get("quality_score")],
        ["launch_readiness", verdict.get("launch_readiness")],
    ]
    _add_sheet(workbook, "Run_Summary", ["field", "value"], summary_rows)
    _add_sheet(
        workbook,
        "Users",
        [
            "run_id",
            "test_user_id",
            "user_type",
            "persona_key",
            "persona_label",
            "source_persona_id",
            "journey_started_at_utc",
            "journey_completed_at_utc",
            "status",
        ],
        persona["users"],
    )
    _add_sheet(
        workbook,
        "Journey_Events",
        [
            "run_id",
            "test_user_id",
            "user_type",
            "persona_key",
            "journey_date_utc",
            "event_timestamp_utc",
            "stage_number",
            "stage_name",
            "user_action",
            "screenshot",
            "status",
        ],
        persona["journeys"],
    )
    _add_sheet(
        workbook,
        "Recommendations",
        [
            "run_id",
            "test_user_id",
            "event_timestamp_utc",
            "journey_stage",
            "request_id",
            "http_status",
            "plate_rank",
            "plate_id",
            "dish_rank",
            "dish_id",
            "dish_name",
            "plate_form",
            "support",
            "plate_score",
        ],
        persona["recommendations"],
    )
    _add_sheet(
        workbook,
        "Final_Meal_Plans",
        [
            "run_id",
            "test_user_id",
            "event_timestamp_utc",
            "journey_stage",
            "request_id",
            "plate_rank",
            "plate_id",
            "plate_form",
            "dish_ids",
            "dish_names",
            "support",
            "plate_score",
            "warnings",
        ],
        persona["plans"],
    )
    _add_sheet(
        workbook,
        "Test_Results",
        [
            "run_id",
            "test_suite",
            "test_user_id",
            "persona_key",
            "test_case",
            "expected_result",
            "actual_result",
            "status",
            "failure_reason",
            "completed_at_utc",
        ],
        persona["tests"],
    )
    _add_sheet(
        workbook,
        "Errors",
        [
            "run_id",
            "test_user_id",
            "timestamp_utc",
            "component",
            "error",
            "evidence_path",
        ],
        persona["errors"],
    )
    _add_sheet(
        workbook,
        "Source_Traceability",
        ["field", "value"],
        [
            ["source_workbook", str(source_workbook)],
            ["source_sha256", _sha256(source_workbook) if source_workbook.exists() else None],
            ["source_persona_sheet", "Persona_Master_v3"],
            ["source_plan_sheet", "Weekly_Class_Plan_v3"],
            ["mapped_source_personas", sorted(source_ids)],
            ["copied_source_persona_rows", len(source_persona_rows)],
            ["copied_source_plan_rows", len(source_plan_rows)],
        ],
    )
    _add_sheet(workbook, "Source_Personas", source_persona_headers, source_persona_rows)
    _add_sheet(workbook, "Source_Meal_Plans", source_plan_headers, source_plan_rows)
    workbook.save(paths.workbook)
    _validate_workbook(paths.workbook, require_users=kind == "persona-journey")
    _write_manifest(report_dir, paths, run_id, kind, environment, source_workbook, generated_at)
    _write_archive(report_dir, paths, kind)
    _json_log(
        "report_built",
        kind=kind,
        run_id=run_id,
        workbook=str(paths.workbook),
        archive=str(paths.archive),
        users=len(persona["users"]),
        tests=len(persona["tests"]),
    )
    return paths


def main() -> int:
    """Parse CLI arguments and publish a validated workbook and ZIP report."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("report_dir", type=Path)
    parser.add_argument("--kind", choices=("persona-journey", "quality-gate"), required=True)
    parser.add_argument("--source-workbook", type=Path, default=DEFAULT_SOURCE_WORKBOOK)
    parser.add_argument("--run-id", default=None)
    parser.add_argument("--environment", default=os.environ.get("GHAR_TEST_ENV", "local"))
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    try:
        paths = build_report(
            args.report_dir,
            kind=args.kind,
            source_workbook=args.source_workbook,
            run_id=args.run_id,
            environment=args.environment,
        )
    except Exception as error:
        error_path = args.report_dir / "errors.log"
        error_path.parent.mkdir(parents=True, exist_ok=True)
        with error_path.open("a", encoding="utf-8") as handle:
            handle.write(
                json.dumps(
                    {
                        "timestamp_utc": datetime.now(UTC).isoformat(),
                        "event": "report_build_failed",
                        "error_type": type(error).__name__,
                        "error": str(error),
                    },
                    sort_keys=True,
                )
                + "\n"
            )
        LOGGER.exception(json.dumps({"event": "report_build_failed", "error": str(error)}))
        return 1
    _json_log("report_ready", workbook=paths.workbook, archive=paths.archive)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
