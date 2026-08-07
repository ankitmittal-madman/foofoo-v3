"""Audit and canonicalize the supplied synthetic Indian household datasets."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

POSITIVE_EVENTS = {
    "cooked": 1.0,
    "locked": 0.95,
    "saved": 0.8,
    "planned": 0.7,
    "rated": 0.6,
    "substituted": 0.45,
    "search": 0.12,
    "viewed": 0.05,
}
NEGATIVE_EVENTS = {"never": -1.0, "skipped": -0.8, "not_today": -0.35}
PREFERENCE_WEIGHTS = {"love": 1.0, "like": 0.7, "neutral": 0.05, "avoid": -0.8, "never": -1.0}
MEMBER_SENTIMENT_WEIGHTS = {
    "love": 1.0,
    "like": 0.65,
    "neutral": 0.05,
    "dislike": -0.65,
    "hate": -1.0,
    "avoid": -1.0,
}
NON_VEG_WORDS = {"chicken", "mutton", "fish", "gosht", "mangsho", "prawn", "egg"}
CUISINE_REGIONS = {
    "andhra": "south",
    "bengali": "east",
    "gujarati": "west",
    "goan": "west",
    "kashmiri": "north",
    "kerala": "south",
    "maharashtrian": "west",
    "malabar": "south",
    "mughlai": "north",
    "odia": "east",
    "punjabi": "north",
    "rajasthani": "north",
    "tamil": "south",
    "telangana": "south",
    "up": "north",
}
ALLERGEN_INGREDIENTS = {
    "peanut": "peanut",
    "groundnut": "peanut",
    "milk": "dairy",
    "curd": "dairy",
    "paneer": "dairy",
    "wheat": "wheat",
    "cashew": "tree_nut",
    "almond": "tree_nut",
    "fish": "fish",
    "prawn": "shellfish",
    "mustard": "mustard",
    "egg": "egg",
}
SPICE_GROUPS = {
    "chilli_forward": {"chilli", "chilli_powder", "green_chilli", "red_chilli"},
    "warming_whole_spices": {
        "cardamom_green",
        "cardamom_black",
        "cinnamon",
        "clove",
        "pepper",
        "star_anise",
    },
    "tempering_spices": {
        "cumin",
        "cumin_seeds",
        "mustard",
        "mustard_seeds",
        "nigella_seeds",
        "fenugreek",
        "asafoetida",
    },
}
NUTRITION_TRAITS = {
    "contains_pulse": {"dal", "lentil", "chana", "chickpea", "rajma", "beans", "moong"},
    "contains_dairy": {"milk", "curd", "paneer", "yogurt"},
    "contains_egg": {"egg"},
    "contains_fish_or_seafood": {"fish", "prawn", "shrimp"},
    "contains_meat": {"chicken", "mutton", "lamb", "goat"},
    "contains_leafy_green": {"spinach", "palak", "methi", "amaranth"},
    "contains_millet": {"millet", "ragi", "jowar", "bajra"},
}


def normalize_name(value: str) -> str:
    value = value.casefold().replace("&", " and ")
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value).split())


def lookup_key(value: str) -> str:
    return " ".join(sorted(token for token in normalize_name(value).split() if token != "and"))


def canonical_id(name: str) -> str:
    token = re.sub(r"[^A-Z0-9]+", "_", normalize_name(name).upper()).strip("_")
    if token:
        return f"DISH_{token[:64]}"
    return "DISH_" + hashlib.sha256(name.encode()).hexdigest()[:16].upper()


def rows(path: Path, sheet: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        iterator = worksheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(iterator)]
        return [
            dict(zip(headers, values, strict=False))
            for values in iterator
            if any(value is not None and value != "" for value in values)
        ]
    finally:
        workbook.close()


def sheet_inventory(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        inventory = []
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            headers = [str(value) for value in next(iterator, ())]
            row_count = sum(
                1
                for values in iterator
                if any(value is not None and value != "" for value in values)
            )
            inventory.append({"sheet": worksheet.title, "rows": row_count, "columns": headers})
        return inventory
    finally:
        workbook.close()


@dataclass(frozen=True)
class AuditFinding:
    severity: str
    dataset: str
    check: str
    count: int
    detail: str


def _missing_fk(
    child_rows: Iterable[dict[str, Any]],
    child_key: str,
    parent_rows: Iterable[dict[str, Any]],
    parent_key: str,
) -> int:
    parents = {row[parent_key] for row in parent_rows}
    return sum(row.get(child_key) not in parents for row in child_rows)


def audit_dataset(path: Path, label: str) -> dict[str, Any]:
    inventory = sheet_inventory(path)
    sheet_names = {entry["sheet"] for entry in inventory}
    required = {
        "DATA_households",
        "DATA_users",
        "DATA_food_preferences",
        "DATA_meal_history",
        "DATA_recommendation_events",
    }
    findings: list[AuditFinding] = []
    missing_sheets = sorted(required - sheet_names)
    if missing_sheets:
        findings.append(
            AuditFinding(
                "error", label, "required_sheets", len(missing_sheets), ",".join(missing_sheets)
            )
        )
        return {
            "dataset": label,
            "path": str(path),
            "inventory": inventory,
            "findings": [asdict(row) for row in findings],
        }

    loaded = {name: rows(path, name) for name in sheet_names if name.startswith("DATA_")}
    for name, values in loaded.items():
        primary_key = next(iter(values[0])) if values else None
        if primary_key:
            ids = [row[primary_key] for row in values]
            duplicates = len(ids) - len(set(ids))
            if duplicates:
                findings.append(
                    AuditFinding("error", label, f"{name}.duplicate_pk", duplicates, primary_key)
                )

    missing_user_households = _missing_fk(
        loaded["DATA_users"], "household_id", loaded["DATA_households"], "household_id"
    )
    missing_history_households = _missing_fk(
        loaded["DATA_meal_history"], "household_id", loaded["DATA_households"], "household_id"
    )
    missing_event_households = _missing_fk(
        loaded["DATA_recommendation_events"],
        "household_id",
        loaded["DATA_households"],
        "household_id",
    )
    for check, count in (
        ("users.household_fk", missing_user_households),
        ("history.household_fk", missing_history_households),
        ("events.household_fk", missing_event_households),
    ):
        if count:
            findings.append(AuditFinding("error", label, check, count, "orphan rows"))

    fk_checks = (
        ("members.household_fk", "DATA_members", "household_id", "DATA_households", "household_id"),
        (
            "preferences.household_fk",
            "DATA_food_preferences",
            "household_id",
            "DATA_households",
            "household_id",
        ),
        (
            "regional.household_fk",
            "DATA_regional_taste",
            "household_id",
            "DATA_households",
            "household_id",
        ),
        (
            "exclusions.household_fk",
            "DATA_exclusions",
            "household_id",
            "DATA_households",
            "household_id",
        ),
        (
            "consumers.meal_fk",
            "DATA_meal_consumers",
            "meal_event_id",
            "DATA_meal_history",
            "meal_event_id",
        ),
        ("events.user_fk", "DATA_recommendation_events", "user_id", "DATA_users", "user_id"),
    )
    for check, child, child_key, parent, parent_key in fk_checks:
        if child not in loaded or parent not in loaded:
            continue
        count = _missing_fk(loaded[child], child_key, loaded[parent], parent_key)
        if count:
            findings.append(AuditFinding("error", label, check, count, "orphan rows"))

    critical_fields = {
        "DATA_households": ("household_id", "current_state_id", "household_size"),
        "DATA_food_preferences": ("household_id", "diet_pattern"),
        "DATA_meal_history": (
            "meal_event_id",
            "household_id",
            "meal_date",
            "meal_slot",
            "dish_raw_name",
        ),
        "DATA_recommendation_events": (
            "event_id",
            "household_id",
            "dish_id",
            "event_type",
            "event_at",
        ),
    }
    for sheet, fields in critical_fields.items():
        for field in fields:
            missing = sum(row.get(field) in (None, "") for row in loaded[sheet])
            if missing:
                findings.append(
                    AuditFinding(
                        "error", label, f"{sheet}.{field}.missing", missing, "required values"
                    )
                )
    known_labels = set(POSITIVE_EVENTS) | set(NEGATIVE_EVENTS) | {"substituted"}
    unknown_labels = Counter(
        str(row.get("event_type") or "").casefold()
        for row in loaded["DATA_recommendation_events"]
        if str(row.get("event_type") or "").casefold() not in known_labels
    )
    if unknown_labels:
        findings.append(
            AuditFinding(
                "warning",
                label,
                "events.unknown_labels",
                sum(unknown_labels.values()),
                json.dumps(dict(sorted(unknown_labels.items()))),
            )
        )
    malformed_context = sum(
        bool(row.get("context_json")) and not _context(row.get("context_json"))
        for row in loaded["DATA_recommendation_events"]
    )
    if malformed_context:
        findings.append(
            AuditFinding("error", label, "events.context_json", malformed_context, "malformed JSON")
        )
    history_dates = Counter(str(row.get("meal_date") or "") for row in loaded["DATA_meal_history"])
    if len(history_dates) < 7:
        findings.append(
            AuditFinding(
                "warning",
                label,
                "history.temporal_coverage",
                len(history_dates),
                "fewer than seven distinct dates; weekday/weekend learning is biased",
            )
        )
    household_count = len(loaded["DATA_households"])
    for sheet in (
        "DATA_members",
        "DATA_regional_taste",
        "DATA_cooking_capability",
        "DATA_festival_seasonal",
    ):
        if sheet in loaded and len(loaded[sheet]) < household_count * 0.5:
            findings.append(
                AuditFinding(
                    "warning",
                    label,
                    f"{sheet}.coverage",
                    len(loaded[sheet]),
                    f"rows for {household_count} households",
                )
            )

    dish_ids = {
        row.get("canonical_dish_id")
        for row in loaded["DATA_meal_history"]
        if row.get("canonical_dish_id")
    }
    event_dishes = {
        row.get("dish_id") for row in loaded["DATA_recommendation_events"] if row.get("dish_id")
    }
    findings.append(
        AuditFinding(
            "info", label, "dish_namespace", len(dish_ids | event_dishes), "unique dish ids"
        )
    )
    findings.append(
        AuditFinding(
            "warning",
            label,
            "synthetic_provenance",
            len(loaded["DATA_meal_history"]) + len(loaded["DATA_recommendation_events"]),
            "AI-generated rows are suitable for pipeline validation, not production-lift claims",
        )
    )
    return {
        "dataset": label,
        "path": str(path),
        "inventory": inventory,
        "findings": [asdict(row) for row in findings],
    }


def _json_list(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        return [str(item) for item in value]
    try:
        parsed = json.loads(str(value))
        return [str(item) for item in parsed] if isinstance(parsed, list) else [str(parsed)]
    except json.JSONDecodeError:
        return [str(value)]


def _state_regions(catalog: dict[str, Any]) -> dict[str, str]:
    return {row["state_id"]: row["region"] for row in catalog["states"]}


def _recipe_metadata(recipe: dict[str, Any]) -> dict[str, Any]:
    ingredients = [str(value).split(" — ", 1)[0] for value in recipe.get("ingredients", [])]
    normalized_ingredients = {normalize_name(value).replace(" ", "_") for value in ingredients}
    diet = str(recipe.get("diet") or "").casefold()
    diet_types = {
        "veg": ["vegetarian"],
        "vegetarian": ["vegetarian"],
        "vegan": ["vegan"],
        "non_veg": ["nonvegetarian"],
        "non-veg": ["nonvegetarian"],
    }.get(diet, [])
    allergens = sorted(
        {
            allergen
            for ingredient in ingredients
            for token, allergen in ALLERGEN_INGREDIENTS.items()
            if token in ingredient.casefold()
        }
    )
    raw_spice_level = recipe.get("spice_level")
    spice_level = None
    if raw_spice_level not in (None, ""):
        spice_level = max(1, min(5, int(raw_spice_level)))
    spice_profiles = [
        group for group, tokens in SPICE_GROUPS.items() if normalized_ingredients & tokens
    ]
    if spice_level is not None:
        spice_profiles.append(
            "mild" if spice_level <= 2 else "medium" if spice_level == 3 else "hot"
        )
    nutrition_traits = [
        trait
        for trait, tokens in NUTRITION_TRAITS.items()
        if any(
            any(token in ingredient for token in tokens) for ingredient in normalized_ingredients
        )
    ]
    attributes = recipe.get("attribute_basis", {})
    return {
        "ingredients": ingredients,
        "diet_types": diet_types,
        "allergens": allergens,
        "spice_level": spice_level,
        "spice_profiles": sorted(set(spice_profiles)),
        "nutrition_traits": sorted(set(nutrition_traits)),
        "dish_categories": sorted(
            str(value) for value in attributes.get("dish_category", []) if value
        ),
        "cooking_methods": sorted(
            str(value) for value in attributes.get("cooking_method", []) if value
        ),
        "cook_minutes": int(recipe.get("total_mins") or 0) or None,
    }


def build_ontology(
    catalog_path: Path, primary_workbook: Path, enrichment_workbook: Path, recipes_path: Path
) -> tuple[dict[str, Any], dict[str, str]]:
    catalog = json.loads(catalog_path.read_text())
    recipes = json.loads(recipes_path.read_text())
    recipes_by_name = {lookup_key(name): value for name, value in recipes.items()}
    dishes: dict[str, dict[str, Any]] = {}
    name_to_id: dict[str, str] = {}

    for source in catalog["dishes"]:
        dish_id = source["dish_id"]
        name = source["dish_name"]
        recipe = recipes_by_name.get(lookup_key(name), {})
        metadata = _recipe_metadata(recipe)
        diets = metadata["diet_types"] or [
            "vegetarian" if tag == "veg" else str(tag) for tag in source.get("diet_tags", [])
        ]
        cuisine = str(recipe.get("cuisine") or "")
        dishes[dish_id] = {
            "id": dish_id,
            "name": name,
            "aliases": [],
            "ingredients": metadata["ingredients"],
            "allergens": metadata["allergens"],
            "diet_types": diets,
            "cuisines": [cuisine] if cuisine else [],
            "regions": [CUISINE_REGIONS[cuisine]] if cuisine in CUISINE_REGIONS else [],
            "observed_regions": [],
            "meal_slots": [str(value).casefold() for value in source.get("slots", [])],
            "cooking_methods": metadata["cooking_methods"],
            "dish_categories": metadata["dish_categories"],
            "spice_level": metadata["spice_level"],
            "spice_profiles": metadata["spice_profiles"],
            "nutrition_traits": metadata["nutrition_traits"],
            "cook_minutes": metadata["cook_minutes"],
            "seasons": [],
            "occasions": [],
            "substitutes": [],
            "source_datasets": ["dataset_1_catalog"],
        }
        name_to_id[lookup_key(name)] = dish_id

    state_region = _state_regions(catalog)
    for workbook, source_name in (
        (primary_workbook, "dataset_1"),
        (enrichment_workbook, "dataset_2"),
    ):
        households = {row["household_id"]: row for row in rows(workbook, "DATA_households")}
        history = rows(workbook, "DATA_meal_history")
        stats: dict[str, dict[str, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))
        raw_names: dict[str, Counter[str]] = defaultdict(Counter)
        for event in history:
            normalized = lookup_key(str(event["dish_raw_name"]))
            dish_id = name_to_id.get(normalized)
            if dish_id is None:
                dish_id = canonical_id(str(event["dish_raw_name"]))
                name_to_id[normalized] = dish_id
            raw_names[dish_id][str(event["dish_raw_name"])] += 1
            stats[dish_id]["slots"][str(event["meal_slot"]).casefold()] += 1
            household = households.get(event["household_id"], {})
            state = str(household.get("current_state_id") or "")
            region = state_region.get(state, state.casefold())
            if region:
                stats[dish_id]["regions"][region] += 1
            if dish_id not in dishes:
                name = str(event["dish_raw_name"])
                recipe = recipes_by_name.get(normalized, {})
                is_non_veg = bool(set(normalized.split()) & NON_VEG_WORDS)
                metadata = _recipe_metadata(recipe)
                cuisine = str(recipe.get("cuisine") or "")
                dishes[dish_id] = {
                    "id": dish_id,
                    "name": name,
                    "aliases": [],
                    "ingredients": metadata["ingredients"],
                    "allergens": metadata["allergens"],
                    "diet_types": metadata["diet_types"]
                    or (["nonvegetarian"] if is_non_veg else ["unknown"]),
                    "cuisines": [cuisine] if cuisine else [],
                    "regions": [CUISINE_REGIONS[cuisine]] if cuisine in CUISINE_REGIONS else [],
                    "observed_regions": [],
                    "meal_slots": [],
                    "cooking_methods": metadata["cooking_methods"],
                    "dish_categories": metadata["dish_categories"],
                    "spice_level": metadata["spice_level"],
                    "spice_profiles": metadata["spice_profiles"],
                    "nutrition_traits": metadata["nutrition_traits"],
                    "cook_minutes": metadata["cook_minutes"],
                    "seasons": [],
                    "occasions": [],
                    "substitutes": [],
                    "source_datasets": [],
                }
            if source_name not in dishes[dish_id]["source_datasets"]:
                dishes[dish_id]["source_datasets"].append(source_name)

        for dish_id, dimensions in stats.items():
            if not dishes[dish_id]["meal_slots"]:
                dishes[dish_id]["meal_slots"] = sorted(dimensions["slots"])
            dishes[dish_id]["observed_regions"] = [
                name for name, _ in dimensions["regions"].most_common(4)
            ]
            if not dishes[dish_id]["regions"]:
                dishes[dish_id]["regions"] = dishes[dish_id]["observed_regions"]
            aliases = [
                name
                for name, _ in raw_names[dish_id].most_common()
                if name != dishes[dish_id]["name"]
            ]
            dishes[dish_id]["aliases"] = aliases

    substitutions: dict[str, Counter[str]] = defaultdict(Counter)
    seasonal: dict[str, set[str]] = defaultdict(set)
    occasions: dict[str, set[str]] = defaultdict(set)
    for workbook in (primary_workbook, enrichment_workbook):
        valid_households = {str(row["household_id"]) for row in rows(workbook, "DATA_households")}
        for event in rows(workbook, "DATA_recommendation_events"):
            if str(event.get("household_id")) not in valid_households:
                continue
            source = str(event.get("dish_id") or "")
            target = str(event.get("substitute_dish_id") or "")
            if source in dishes and target in dishes and source != target:
                substitutions[source][target] += 1
        for event in rows(workbook, "DATA_festival_seasonal"):
            if str(event.get("household_id")) not in valid_households:
                continue
            occasion = str(event.get("occasion_id") or "").casefold()
            for dish_id in _json_list(event.get("dish_ids")):
                if dish_id not in dishes:
                    continue
                if occasion.startswith("season_"):
                    seasonal[dish_id].add(occasion.removeprefix("season_"))
                elif occasion:
                    occasions[dish_id].add(occasion.removeprefix("festival_"))
    for dish_id, dish in dishes.items():
        dish["substitutes"] = [target for target, _ in substitutions[dish_id].most_common(8)]
        dish["seasons"] = sorted(seasonal[dish_id])
        dish["occasions"] = sorted(occasions[dish_id])

    ordered = sorted(dishes.values(), key=lambda row: row["id"])
    nodes = []
    relations = []
    seen_nodes: set[str] = set()
    for dish in ordered:
        nodes.append({"id": dish["id"], "type": "dish", "name": dish["name"]})
        seen_nodes.add(dish["id"])
        relation_dimensions = (
            ("ingredients", "ingredient", "contains"),
            ("allergens", "allergy", "incompatible_with"),
            ("cuisines", "cuisine", "belongs_to"),
            ("regions", "region", "eaten_in"),
            ("meal_slots", "meal_slot", "served_at"),
            ("diet_types", "diet_type", "compatible_with"),
            ("cooking_methods", "cooking_technique", "cooked_with"),
            ("dish_categories", "meal_class", "belongs_to"),
            ("spice_profiles", "spice_profile", "has_spice_profile"),
            ("nutrition_traits", "nutrition_trait", "has_nutrition_trait"),
            ("seasons", "season", "eaten_in"),
            ("occasions", "occasion", "served_at"),
        )
        for field, node_type, relation in relation_dimensions:
            for value in dish[field]:
                target = f"{node_type.upper()}_{canonical_id(str(value)).removeprefix('DISH_')}"
                if target not in seen_nodes:
                    nodes.append({"id": target, "type": node_type, "name": value})
                    seen_nodes.add(target)
                relations.append({"source": dish["id"], "relation": relation, "target": target})
        for target in dish["substitutes"]:
            relations.append(
                {"source": dish["id"], "relation": "substitutes_for", "target": target}
            )
    for dish in ordered:
        left = set(dish["ingredients"])
        left_context = set(dish["meal_slots"] + dish["dish_categories"] + dish["cuisines"])
        ranked_similar = []
        for other in ordered:
            if other["id"] == dish["id"]:
                continue
            right = set(other["ingredients"])
            right_context = set(other["meal_slots"] + other["dish_categories"] + other["cuisines"])
            ingredient_score = len(left & right) / max(1, len(left | right))
            context_score = len(left_context & right_context) / max(
                1, len(left_context | right_context)
            )
            score = 0.7 * ingredient_score + 0.3 * context_score
            if score > 0:
                ranked_similar.append((score, other["id"]))
        for _, target in sorted(ranked_similar, key=lambda row: (-row[0], row[1]))[:3]:
            relations.append({"source": dish["id"], "relation": "similar_to", "target": target})
    return {
        "version": "indian-food-ontology-v2",
        "generated_at": datetime.now(UTC).isoformat(),
        "nodes": nodes,
        "relations": relations,
        "dishes": ordered,
    }, name_to_id


def _event_weight(event: dict[str, Any]) -> float:
    event_type = str(event.get("event_type") or "").casefold()
    if event_type == "rated" and event.get("feedback_score") is not None:
        return (float(event["feedback_score"]) - 3.0) / 2.0
    if event_type == "substituted":
        return -0.4
    return POSITIVE_EVENTS.get(event_type, NEGATIVE_EVENTS.get(event_type, 0.1))


def _context(value: Any) -> dict[str, Any]:
    if not value:
        return {}
    try:
        parsed = json.loads(str(value))
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def _day_type(timestamp: str) -> str:
    try:
        day = datetime.fromisoformat(timestamp[:10]).weekday()
    except ValueError:
        return "unknown"
    return "weekend" if day >= 5 else "weekday"


def _ordinal_level(value: Any) -> int:
    normalized = str(value or "0").strip().casefold()
    aliases = {"low": 1, "mild": 1, "medium": 3, "moderate": 3, "high": 5, "hot": 5}
    if normalized in aliases:
        return aliases[normalized]
    try:
        return max(0, min(5, int(float(normalized))))
    except ValueError:
        return 0


def build_interactions(
    workbook: Path, dataset_name: str, name_to_id: dict[str, str]
) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    prefix = f"{dataset_name}:"
    valid_households = {str(row["household_id"]) for row in rows(workbook, "DATA_households")}
    for event in rows(workbook, "DATA_recommendation_events"):
        if str(event["household_id"]) not in valid_households:
            continue
        dish_id = str(event["dish_id"])
        timestamp = str(event.get("event_at") or "")
        context = _context(event.get("context_json"))
        output.append(
            {
                "event_id": prefix + str(event.get("event_id") or ""),
                "household_id": prefix + str(event["household_id"]),
                "dish_id": dish_id,
                "timestamp": timestamp,
                "meal_slot": str(event.get("meal_slot") or "").casefold(),
                "weight": _event_weight(event),
                "event_type": str(event.get("event_type") or ""),
                "member_id": None,
                "day_type": str(context.get("day_type") or _day_type(timestamp)).casefold(),
                "context": context,
                "source_dataset": dataset_name,
            }
        )
        substitute = str(event.get("substitute_dish_id") or "")
        if substitute:
            output.append(
                {
                    "event_id": prefix + str(event.get("event_id") or "") + ":substitute",
                    "household_id": prefix + str(event["household_id"]),
                    "dish_id": substitute,
                    "timestamp": timestamp,
                    "meal_slot": str(event.get("meal_slot") or "").casefold(),
                    "weight": 0.75,
                    "event_type": "substitute_selected",
                    "member_id": None,
                    "day_type": str(context.get("day_type") or _day_type(timestamp)).casefold(),
                    "context": context,
                    "source_dataset": dataset_name,
                }
            )
    history = rows(workbook, "DATA_meal_history")
    history_by_event: dict[str, dict[str, Any]] = {}
    for event in history:
        dish_id = name_to_id.get(
            lookup_key(str(event["dish_raw_name"])), str(event["canonical_dish_id"])
        )
        satisfaction = float(event.get("satisfaction_score") or 3)
        satisfaction_weight = (satisfaction - 3.0) / 2.0
        sentiment_weight = MEMBER_SENTIMENT_WEIGHTS.get(
            str(event.get("sentiment") or "").casefold(), 0.0
        )
        repeat_weight = {
            "very_high": 0.4,
            "high": 0.25,
            "medium": 0.0,
            "low": -0.25,
            "never": -0.5,
        }.get(str(event.get("repeat_desire") or "").casefold(), 0.0)
        weight = max(
            -1.0,
            min(1.0, 0.6 * satisfaction_weight + 0.25 * sentiment_weight + 0.15 * repeat_weight),
        )
        timestamp = str(event.get("meal_date") or "")
        context_flags = [value.casefold() for value in _json_list(event.get("context_flags"))]
        row = {
            "event_id": prefix + str(event.get("meal_event_id") or ""),
            "household_id": prefix + str(event["household_id"]),
            "dish_id": dish_id,
            "timestamp": timestamp,
            "meal_slot": str(event.get("meal_slot") or "").casefold(),
            "weight": weight,
            "event_type": "meal_history",
            "member_id": None,
            "day_type": "weekend" if "weekend" in context_flags else _day_type(timestamp),
            "context": {
                "flags": context_flags,
                "source_mode": str(event.get("source_mode") or "").casefold(),
                "leftover_level": str(event.get("leftover_level") or "").casefold(),
                "portion_size": str(event.get("portion_size") or "").casefold(),
            },
            "source_dataset": dataset_name,
        }
        output.append(row)
        history_by_event[str(event.get("meal_event_id"))] = row

    for event in rows(workbook, "DATA_meal_consumers"):
        meal = history_by_event.get(str(event.get("meal_event_id")))
        if meal is None:
            continue
        sentiment = str(event.get("member_sentiment") or "").casefold()
        weight = MEMBER_SENTIMENT_WEIGHTS.get(sentiment, 0.0)
        output.append(
            {
                **meal,
                "event_id": prefix + str(event.get("meal_consumer_id") or ""),
                "weight": weight,
                "event_type": "member_meal_feedback",
                "member_id": prefix + str(event.get("member_id") or ""),
                "context": {
                    **meal["context"],
                    "member_sentiment": sentiment,
                    "finished_portion": str(event.get("finished_portion") or "").casefold(),
                },
            }
        )

    for event in rows(workbook, "DATA_dish_preferences"):
        household_id = str(event.get("household_id") or "")
        if household_id not in valid_households:
            continue
        preference_type = str(event.get("preference_type") or "neutral").casefold()
        confidence = float(event.get("confidence_score") or 50) / 100.0
        confirmed_boost = 1.0 if bool(event.get("confirmed_by_user")) else 0.85
        weight = PREFERENCE_WEIGHTS.get(preference_type, 0.0) * confidence * confirmed_boost
        output.append(
            {
                "event_id": prefix + str(event.get("dish_preference_id") or ""),
                "household_id": prefix + household_id,
                "dish_id": str(event.get("dish_id") or ""),
                "timestamp": "2026-01-01T00:00:00Z",
                "meal_slot": "",
                "weight": max(-1.0, min(1.0, weight)),
                "event_type": f"dish_preference:{preference_type}",
                "member_id": prefix + str(event.get("member_id") or ""),
                "day_type": "profile",
                "context": {
                    "evidence_source": str(event.get("evidence_source") or "").casefold(),
                    "confirmed_by_user": bool(event.get("confirmed_by_user")),
                },
                "source_dataset": dataset_name,
            }
        )
    return sorted(
        output,
        key=lambda row: (
            row["household_id"],
            row["timestamp"],
            row["dish_id"],
            row["event_id"],
        ),
    )


def build_household_features(workbook: Path, dataset_name: str) -> list[dict[str, Any]]:
    prefix = f"{dataset_name}:"
    preferences: dict[str, set[str]] = defaultdict(set)
    for row in rows(workbook, "DATA_food_preferences"):
        household_id = str(row["household_id"])
        preferences[household_id].update(
            {
                f"diet:{str(row.get('diet_pattern') or 'unknown').casefold()}",
                f"spice:{_ordinal_level(row.get('spice_level'))}",
                f"sweetness:{_ordinal_level(row.get('sweetness_level'))}",
                f"oil:{_ordinal_level(row.get('oil_level'))}",
                f"texture:{str(row.get('texture_preference') or 'unknown').casefold()}",
                f"home_cooked_days:{min(7, int(row.get('home_cooked_days_week') or 0))}",
            }
        )
        preferences[household_id].update(
            f"fasting:{value.casefold()}" for value in _json_list(row.get("fasting_patterns"))
        )
        preferences[household_id].update(
            f"comfort:{value}" for value in _json_list(row.get("comfort_dish_ids"))
        )
    regional: dict[str, set[str]] = defaultdict(set)
    for row in rows(workbook, "DATA_regional_taste"):
        household_id = str(row["household_id"])
        regional[household_id].update(
            {
                f"region:{str(row.get('cuisine_region_id') or 'unknown').casefold()}",
                f"regional_familiarity:{str(row.get('familiarity') or 'unknown').casefold()}",
                "regional_authenticity:"
                + str(row.get("authenticity_preference") or "unknown").casefold(),
                f"regional_score:{int(float(row.get('preference_score') or 0) // 20)}",
            }
        )
    capability: dict[str, set[str]] = defaultdict(set)
    for row in rows(workbook, "DATA_cooking_capability"):
        household_id = str(row["household_id"])
        capability[household_id].update(
            {
                f"cook_role:{str(row.get('primary_cook_role') or 'unknown').casefold()}",
                f"cook_frequency:{str(row.get('cook_frequency') or 'unknown').casefold()}",
                f"skill:{str(row.get('skill_level') or 'unknown').casefold()}",
                f"complexity:{str(row.get('preferred_complexity') or 'unknown').casefold()}",
                f"weekday_minutes:{int(row.get('weekday_minutes') or 0) // 15 * 15}",
                f"weekend_minutes:{int(row.get('weekend_minutes') or 0) // 30 * 30}",
                f"novelty:{int(row.get('novelty_willingness') or 0)}",
            }
        )
        capability[household_id].update(
            f"equipment:{value.casefold()}" for value in _json_list(row.get("equipment_ids"))
        )
    exclusions: dict[str, set[str]] = defaultdict(set)
    for row in rows(workbook, "DATA_exclusions"):
        if not bool(row.get("confirmed_by_user")):
            continue
        exclusions[str(row["household_id"])].add(
            "exclude:"
            + str(row.get("exclusion_type") or "unknown").casefold()
            + ":"
            + str(row.get("entity_id") or "unknown").casefold()
        )
    health: dict[str, set[str]] = defaultdict(set)
    for row in rows(workbook, "DATA_health_goals"):
        household_id = str(row["household_id"])
        health[household_id].add(f"health_goal:{str(row.get('goal_code') or 'unknown').casefold()}")
    history: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows(workbook, "DATA_meal_history"):
        household_id = str(row["household_id"])
        timestamp = str(row.get("meal_date") or "")
        flags = [value.casefold() for value in _json_list(row.get("context_flags"))]
        history[household_id][f"history_slot:{str(row.get('meal_slot') or '').casefold()}"] += 1
        history[household_id][f"history_day:{_day_type(timestamp)}"] += 1
        history[household_id][
            f"history_source:{str(row.get('source_mode') or 'unknown').casefold()}"
        ] += 1
        history[household_id][
            f"leftover:{str(row.get('leftover_level') or 'unknown').casefold()}"
        ] += 1
        for flag in flags:
            history[household_id][f"context:{flag}"] += 1
    output = []
    for row in rows(workbook, "DATA_households"):
        household_id = str(row["household_id"])
        history_features = {feature for feature, _ in history[household_id].most_common(12)}
        output.append(
            {
                "household_id": prefix + household_id,
                "features": sorted(
                    {
                        f"state:{str(row.get('current_state_id') or 'unknown').casefold()}",
                        f"origin:{str(row.get('origin_state_ids') or 'unknown').casefold()}",
                        f"setup:{str(row.get('living_setup') or 'unknown').casefold()}",
                        f"size:{min(8, int(row.get('household_size') or 1))}",
                        f"adults:{min(6, int(row.get('adult_count') or 0))}",
                        f"children:{min(4, int(row.get('child_count') or 0))}",
                        f"elders:{min(4, int(row.get('elder_count') or 0))}",
                        f"infants:{min(2, int(row.get('infant_count') or 0))}",
                        "decision_model:" + str(row.get("decision_model") or "unknown").casefold(),
                        *preferences.get(household_id, {"diet:unknown"}),
                        *regional.get(household_id, set()),
                        *capability.get(household_id, set()),
                        *exclusions.get(household_id, set()),
                        *health.get(household_id, set()),
                        *history_features,
                    }
                ),
            }
        )
    return output


def build_weekly_signals(workbook: Path, dataset_name: str) -> list[dict[str, Any]]:
    prefix = f"{dataset_name}:"
    signals: dict[str, dict[str, Counter[str]]] = defaultdict(
        lambda: {
            "slot": Counter(),
            "day_type": Counter(),
            "day_slot_dish": Counter(),
            "source": Counter(),
            "leftover": Counter(),
            "context": Counter(),
        }
    )
    for row in rows(workbook, "DATA_meal_history"):
        household_id = str(row["household_id"])
        slot = str(row.get("meal_slot") or "unknown").casefold()
        timestamp = str(row.get("meal_date") or "")
        flags = [value.casefold() for value in _json_list(row.get("context_flags"))]
        day_type = "weekend" if "weekend" in flags else _day_type(timestamp)
        dish_id = str(row.get("canonical_dish_id") or "")
        signals[household_id]["slot"][slot] += 1
        signals[household_id]["day_type"][day_type] += 1
        signals[household_id]["day_slot_dish"][f"{day_type}:{slot}:{dish_id}"] += 1
        signals[household_id]["source"][str(row.get("source_mode") or "unknown").casefold()] += 1
        signals[household_id]["leftover"][
            str(row.get("leftover_level") or "unknown").casefold()
        ] += 1
        signals[household_id]["context"].update(flags)
    capability = {
        str(row["household_id"]): row for row in rows(workbook, "DATA_cooking_capability")
    }
    output = []
    for household_id in sorted(signals):
        values = signals[household_id]
        cooking = capability.get(household_id, {})
        output.append(
            {
                "household_id": prefix + household_id,
                "slot_counts": dict(values["slot"]),
                "day_type_counts": dict(values["day_type"]),
                "top_day_slot_dishes": [
                    {"key": key, "count": count}
                    for key, count in values["day_slot_dish"].most_common(12)
                ],
                "source_counts": dict(values["source"]),
                "leftover_counts": dict(values["leftover"]),
                "context_counts": dict(values["context"]),
                "weekday_minutes": int(cooking.get("weekday_minutes") or 0),
                "weekend_minutes": int(cooking.get("weekend_minutes") or 0),
            }
        )
    return output


def build_household_preference_graph(workbook: Path, dataset_name: str) -> list[dict[str, Any]]:
    prefix = f"{dataset_name}:"
    valid_households = {str(row["household_id"]) for row in rows(workbook, "DATA_households")}
    edges = []
    for row in rows(workbook, "DATA_dish_preferences"):
        household_id = str(row.get("household_id") or "")
        if household_id not in valid_households:
            continue
        preference = str(row.get("preference_type") or "neutral").casefold()
        relation = "avoided_by" if preference in {"avoid", "never"} else "preferred_by"
        edges.append(
            {
                "source": str(row.get("dish_id") or ""),
                "relation": relation,
                "target": prefix + str(row.get("member_id") or household_id),
                "household_id": prefix + household_id,
                "weight": PREFERENCE_WEIGHTS.get(preference, 0.0),
                "confirmed": bool(row.get("confirmed_by_user")),
            }
        )
    history = {str(row.get("meal_event_id")): row for row in rows(workbook, "DATA_meal_history")}
    for row in rows(workbook, "DATA_meal_consumers"):
        meal = history.get(str(row.get("meal_event_id")))
        if meal is None:
            continue
        household_id = str(meal.get("household_id") or "")
        edges.append(
            {
                "source": str(meal.get("canonical_dish_id") or ""),
                "relation": "consumed_by",
                "target": prefix + str(row.get("member_id") or ""),
                "household_id": prefix + household_id,
                "weight": MEMBER_SENTIMENT_WEIGHTS.get(
                    str(row.get("member_sentiment") or "neutral").casefold(), 0.0
                ),
                "confirmed": False,
            }
        )
    return sorted(
        edges,
        key=lambda row: (row["household_id"], row["source"], row["relation"], row["target"]),
    )


def split_interactions(
    interactions: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    by_household: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in interactions:
        if float(row["weight"]) > 0:
            by_household[row["household_id"]].append(row)
    validation_ids: set[str] = set()
    test_ids: set[str] = set()
    for values in by_household.values():
        values.sort(key=lambda row: (row["timestamp"], row["event_id"]))
        distinct_reversed = []
        seen_dishes: set[str] = set()
        for row in reversed(values):
            if row["dish_id"] not in seen_dishes:
                distinct_reversed.append(row)
                seen_dishes.add(row["dish_id"])
        if len(distinct_reversed) >= 3:
            test_ids.add(distinct_reversed[0]["event_id"])
            validation_ids.add(distinct_reversed[1]["event_id"])
    training = [
        row
        for row in interactions
        if row["event_id"] not in validation_ids and row["event_id"] not in test_ids
    ]
    validation = [row for row in interactions if row["event_id"] in validation_ids]
    test = [row for row in interactions if row["event_id"] in test_ids]
    return training, validation, test


SCHEMA_MAP = {
    "dish": {
        "sources": ["catalogs.dishes", "DATA_meal_history", "recipes_v1.json"],
        "fields": [
            "id",
            "name",
            "aliases",
            "ingredients",
            "allergens",
            "diet_types",
            "cuisines",
            "regions",
            "meal_slots",
            "dish_categories",
            "spice_profiles",
            "nutrition_traits",
            "seasons",
            "occasions",
            "substitutes",
        ],
    },
    "household": {
        "sources": [
            "DATA_households",
            "DATA_food_preferences",
            "DATA_regional_taste",
            "DATA_cooking_capability",
            "DATA_exclusions",
            "DATA_health_goals",
            "DATA_meal_history",
        ],
        "fields": ["household_id", "features"],
    },
    "interaction": {
        "sources": [
            "DATA_recommendation_events",
            "DATA_meal_history",
            "DATA_meal_consumers",
            "DATA_dish_preferences",
        ],
        "fields": [
            "event_id",
            "household_id",
            "member_id",
            "dish_id",
            "timestamp",
            "meal_slot",
            "day_type",
            "weight",
            "event_type",
            "context",
        ],
    },
}


def write_artifacts(
    output_dir: Path,
    audits: list[dict[str, Any]],
    ontology: dict[str, Any],
    interactions: list[dict[str, Any]],
    households: list[dict[str, Any]],
    weekly_signals: list[dict[str, Any]],
    preference_graph: list[dict[str, Any]],
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "dataset_audit.json").write_text(
        json.dumps({"datasets": audits}, indent=2, default=str) + "\n"
    )
    (output_dir / "canonical_food_ontology.json").write_text(
        json.dumps(ontology, indent=2, default=str) + "\n"
    )
    (output_dir / "interactions.jsonl").write_text(
        "".join(json.dumps(row, default=str) + "\n" for row in interactions)
    )
    (output_dir / "household_features.jsonl").write_text(
        "".join(json.dumps(row) + "\n" for row in households)
    )
    (output_dir / "weekly_signals.jsonl").write_text(
        "".join(json.dumps(row) + "\n" for row in weekly_signals)
    )
    (output_dir / "household_preference_graph.jsonl").write_text(
        "".join(json.dumps(row) + "\n" for row in preference_graph)
    )
    (output_dir / "schema_map.json").write_text(json.dumps(SCHEMA_MAP, indent=2) + "\n")
    train, validation, test = split_interactions(interactions)
    for name, values in (
        ("interactions_train.jsonl", train),
        ("interactions_validation.jsonl", validation),
        ("interactions_test.jsonl", test),
    ):
        (output_dir / name).write_text(
            "".join(json.dumps(row, default=str) + "\n" for row in values)
        )
    manifest: dict[str, Any] = {
        "version": "foofoo-training-v1",
        "generated_at": datetime.now(UTC).isoformat(),
        "synthetic_only": True,
        "ontology_dishes": len(ontology["dishes"]),
        "interactions": len(interactions),
        "positive_interactions": sum(float(row["weight"]) > 0 for row in interactions),
        "negative_interactions": sum(float(row["weight"]) < 0 for row in interactions),
        "households": len(households),
        "weekly_signal_households": len(weekly_signals),
        "household_graph_edges": len(preference_graph),
        "training_interactions": len(train),
        "validation_interactions": len(validation),
        "test_interactions": len(test),
        "sha256": {},
    }
    for name in (
        "dataset_audit.json",
        "canonical_food_ontology.json",
        "interactions.jsonl",
        "household_features.jsonl",
        "weekly_signals.jsonl",
        "household_preference_graph.jsonl",
        "schema_map.json",
        "interactions_train.jsonl",
        "interactions_validation.jsonl",
        "interactions_test.jsonl",
    ):
        manifest["sha256"][name] = hashlib.sha256((output_dir / name).read_bytes()).hexdigest()
    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit and canonicalize FooFoo training datasets")
    parser.add_argument("--dataset-1", type=Path, required=True)
    parser.add_argument("--dataset-2", type=Path, required=True)
    parser.add_argument("--catalog", type=Path, required=True)
    parser.add_argument("--recipes", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    audits = [
        audit_dataset(args.dataset_1, "dataset_1"),
        audit_dataset(args.dataset_2, "dataset_2"),
    ]
    ontology, name_to_id = build_ontology(
        args.catalog, args.dataset_1, args.dataset_2, args.recipes
    )
    interactions = build_interactions(args.dataset_1, "dataset_1", name_to_id)
    interactions.extend(build_interactions(args.dataset_2, "dataset_2", name_to_id))
    households = build_household_features(args.dataset_1, "dataset_1")
    households.extend(build_household_features(args.dataset_2, "dataset_2"))
    weekly_signals = build_weekly_signals(args.dataset_1, "dataset_1")
    weekly_signals.extend(build_weekly_signals(args.dataset_2, "dataset_2"))
    preference_graph = build_household_preference_graph(args.dataset_1, "dataset_1")
    preference_graph.extend(build_household_preference_graph(args.dataset_2, "dataset_2"))
    write_artifacts(
        args.output_dir,
        audits,
        ontology,
        interactions,
        households,
        weekly_signals,
        preference_graph,
    )


if __name__ == "__main__":
    main()
