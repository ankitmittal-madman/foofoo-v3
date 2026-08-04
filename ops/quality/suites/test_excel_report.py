"""Tests for the Excel/ZIP evidence exporter used by both quality workflows."""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ops.quality.runner.excel_report import REQUIRED_SHEETS, build_report


def _source_workbook(path: Path) -> None:
    """Create the smallest canonical-source-shaped workbook needed by the exporter."""
    workbook = Workbook()
    personas = workbook.active
    personas.title = "Persona_Master_v3"
    personas.append(["persona_id", "persona_name"])
    personas.append(["P01", "Source persona"])
    plans = workbook.create_sheet("Weekly_Class_Plan_v3")
    plans.append(["plan_day_id", "persona_id", "day_of_week", "dinner_primary_class"])
    plans.append(["plan-1", "P01", "Mon", "DN_ONE_POT_DINNER"])
    workbook.save(path)


def _persona_evidence(report_dir: Path) -> None:
    """Create one journey with a stage-timestamped dish recommendation response."""
    persona_dir = report_dir / "personas" / "real-p01"
    persona_dir.mkdir(parents=True)
    (persona_dir / "summary.json").write_text(
        json.dumps(
            {
                "key": "real_p01",
                "label": "Source persona",
                "test_user_id": "test-user-001",
                "user_type": "synthetic",
                "source_persona_id": "P01",
                "ok": True,
                "expect_status": 200,
                "recommendations_status": 200,
                "feature_results": [
                    {
                        "name": "cold-start calibration renders",
                        "status": "pass",
                        "started_at_utc": "2026-08-04T10:03:00+00:00",
                        "completed_at_utc": "2026-08-04T10:04:00+00:00",
                    }
                ],
                "started_at_utc": "2026-08-04T10:00:00+00:00",
                "completed_at_utc": "2026-08-04T10:05:00+00:00",
                "steps": [
                    {
                        "label": "post-onboarding-landing",
                        "timestamp_utc": "2026-08-04T10:04:00+00:00",
                        "screenshot": "001.png",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    (persona_dir / "001.png").write_bytes(b"test screenshot evidence")
    (persona_dir / "recommendation_events.json").write_text(
        json.dumps(
            [
                {
                    "timestamp_utc": "2026-08-04T10:04:10+00:00",
                    "stage_label": "cold-start-loaded",
                    "method": "POST",
                    "endpoint": "/v1/plan",
                    "request_body": {"surface": "calibration"},
                    "response": {
                        "status": 200,
                        "body": {
                            "kind": "calibration",
                            "request_id": "request-1",
                            "slots": {
                                "breakfast": [
                                    {
                                        "name": "Dal",
                                        "cuisine": "North Indian",
                                        "diet": "veg",
                                        "score": 8.2,
                                        "cell_role": "expected_positive",
                                    }
                                ]
                            },
                        },
                    },
                },
                {
                    "timestamp_utc": "2026-08-04T10:04:30+00:00",
                    "stage_label": "post-onboarding-landing",
                    "response": {
                        "status": 200,
                        "body": {
                            "request_id": "request-1",
                            "warnings": [],
                            "plates": [
                                {
                                    "plate_id": "plate-1",
                                    "form": "pair",
                                    "hero_dish_ids": ["dish-1", "dish-2"],
                                    "hero_dish_names": ["Dal", "Rice"],
                                    "support": None,
                                    "plate_score": 9.1,
                                }
                            ],
                        },
                    },
                }
            ]
        ),
        encoding="utf-8",
    )
    (persona_dir / "api_events.json").write_text(
        json.dumps(
            [
                {
                    "timestamp_utc": "2026-08-04T10:04:10+00:00",
                    "stage_label": "cold-start-loaded",
                    "method": "POST",
                    "endpoint": "/v1/plan",
                    "request_body": {"surface": "calibration"},
                    "response": {
                        "status": 200,
                        "body": {
                            "kind": "calibration",
                            "request_id": "request-1",
                            "slots": {
                                "breakfast": [
                                    {
                                        "name": "Dal",
                                        "cuisine": "North Indian",
                                        "diet": "veg",
                                        "score": 8.2,
                                        "cell_role": "expected_positive",
                                    }
                                ]
                            },
                        },
                    },
                },
                {
                    "timestamp_utc": "2026-08-04T10:04:15+00:00",
                    "stage_label": "cold-start-like-breakfast",
                    "method": "POST",
                    "endpoint": "/v1/feedback",
                    "request_body": {
                        "event_type": "like",
                        "dish_name": "Dal",
                        "request_id": "request-1",
                    },
                    "response": {
                        "status": 200,
                        "body": {"id": "feedback-1", "event_type": "like"},
                    },
                }
            ]
        ),
        encoding="utf-8",
    )
    (report_dir / "persona_journeys_result.json").write_text(
        json.dumps(
            {
                "status": "pass",
                "total": 1,
                "ok": 1,
                "failed": 0,
            }
        ),
        encoding="utf-8",
    )


def test_persona_report_contains_journey_dishes_source_plan_and_zip(tmp_path: Path) -> None:
    """A persona run becomes a validated workbook plus a directly downloadable ZIP."""
    report_dir = tmp_path / "report"
    report_dir.mkdir()
    source = tmp_path / "source.xlsx"
    _source_workbook(source)
    _persona_evidence(report_dir)

    paths = build_report(
        report_dir,
        kind="persona-journey",
        source_workbook=source,
        run_id="run-123",
        environment="test",
    )

    workbook = load_workbook(paths.workbook, read_only=True, data_only=True)
    assert tuple(workbook.sheetnames) == REQUIRED_SHEETS
    assert workbook["Users"].max_row == 2
    assert workbook["Journey_Events"].max_row == 2
    assert workbook["Feature_Coverage"].max_row == 2
    assert workbook["Feature_Coverage"]["D2"].value == "cold-start calibration renders"
    assert workbook["API_Events"].max_row == 3
    assert workbook["API_Events"]["H3"].value == "like"
    assert workbook["API_Events"]["I3"].value == "Dal"
    assert workbook["Surface_Dishes"].max_row == 2
    assert workbook["Surface_Dishes"]["H2"].value == "Dal"
    assert workbook["Recommendations"].max_row == 3
    assert workbook["Final_Meal_Plans"].max_row == 2
    assert workbook["Source_Personas"]["A2"].value == "P01"
    assert workbook["Source_Meal_Plans"]["B2"].value == "P01"
    workbook.close()

    with zipfile.ZipFile(paths.archive) as archive:
        names = set(archive.namelist())
    assert "test-results.xlsx" in names
    assert "run-summary.json" in names
    assert "ui-artifacts/personas/real-p01/summary.json" in names


def test_quality_report_exports_generic_test_cases_without_users(tmp_path: Path) -> None:
    """The non-UI quality gate records every generic JUnit case even when no users ran."""
    report_dir = tmp_path / "quality"
    report_dir.mkdir()
    source = tmp_path / "source.xlsx"
    _source_workbook(source)
    (report_dir / "test_results.json").write_text(
        json.dumps(
            {
                "meta": {"generated_at": "2026-08-04T12:00:00+00:00"},
                "verdict": {"quality_score": 100, "launch_readiness": "READY"},
                "steps": [
                    {
                        "name": "unit-core",
                        "status": "pass",
                        "cases": [
                            {"test": "test_module::test_one", "status": "pass"},
                        ],
                        "failures": [],
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    paths = build_report(
        report_dir,
        kind="quality-gate",
        source_workbook=source,
        run_id="run-456",
        environment="test",
    )

    workbook = load_workbook(paths.workbook, read_only=True, data_only=True)
    assert workbook["Users"].max_row == 1
    assert workbook["Test_Results"].max_row == 2
    assert workbook["Test_Results"]["B2"].value == "unit-core"
    workbook.close()
